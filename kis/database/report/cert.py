# -*- coding: utf-8 -*-
"""This module allows querying information about identified certificates."""

__author__ = "Lukas Reiter"
__license__ = "GPL v3.0"
__copyright__ = """Copyright 2022 Lukas Reiter

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""
__version__ = 0.1

import argparse
from openpyxl import Workbook
from typing import List
from database.model import CertType
from database.model import CertInfo
from database.model import ScopeType
from database.model import ServiceState
from database.model import DnsResourceRecordType
from collectors.os.modules.http.core import HttpServiceDescriptor
from database.report.core import BaseReport
from database.report.core import ReportLanguage


class ReportClass(BaseReport):
    """
    this module allows querying information about identified certificates
    """

    def __init__(self, **kwargs) -> None:
        super().__init__(name="cert info",
                         title="Overview Certificates",
                         description="The table provides an overview of all identified certificates.",
                         **kwargs)

    def _filter(self, cert_info: CertInfo) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return cert_info.is_processable(included_items=self._included_items,
                                        excluded_items=self._excluded_items,
                                        scope=self.scope)

    @staticmethod
    def get_add_argparse_arguments(parser_cert: argparse.ArgumentParser):
        """
        This method adds the report's specific command line arguments.
        """
        # setup cert parser
        parser_cert.add_argument("-w", "--workspaces",
                                 metavar="WORKSPACE",
                                 help="query the given workspaces",
                                 nargs="+",
                                 type=str)
        parser_cert.add_argument('--csv',
                                 default=True,
                                 action='store_true',
                                 help='returns gathered information in csv format')
        parser_cert.add_argument('--filter', metavar='IP|NETWORK|DOMAIN|HOSTNAME', type=str, nargs='*',
                                 help='list of IP addresses, IP networks, second-level domains (e.g., megacorpone.com), or '
                                      'host names (e.g., www.megacorpone.com) whose information shall be returned.'
                                      'per default, mentioned items are excluded. add + in front of each item '
                                      '(e.g., +192.168.0.1) to return only these items')
        parser_cert.add_argument('--scope', choices=[item.name for item in ScopeType],
                                 help='return only information about in scope (within) or out of scope (outside) items. '
                                      'per default, all information is returned')

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        descriptor = HttpServiceDescriptor()
        result = [["Workspace",
                   "Type",
                   "Network (NW)",
                   "Scope (NW)",
                   "Company (NW)",
                   "IP Address (IP)",
                   "Version (IP)",
                   "Private IP",
                   "In Scope (IP)",
                   "OS Family",
                   "OS Details",
                   "Second-Level Domain (SLD)",
                   "Scope (SLD)",
                   "Host Name (HN)",
                   "In Scope (HN)",
                   "Name (HN)",
                   "Company (HN)",
                   "Environment",
                   "UDP/TCP",
                   "Port",
                   "Service (SRV)",
                   "Nmap Name (SRV)",
                   "Confidence (SRV)",
                   "State (SRV)",
                   "Reason State",
                   "Banner Information",
                   "Is HTTP",
                   "URL",
                   "HN Coverage",
                   "Common Name",
                   "Issuer",
                   "Invalid CA (Self-Signed)",
                   "Public Key Algorithm",
                   "Key Length",
                   "Public Key Summary",
                   "Proper Key Length",
                   "Hash Algorithm",
                   "Weak Signature",
                   "Cert. Type",
                   "Valid From",
                   "Valid Until",
                   "Valid Years",
                   "Is Valid",
                   "Within Recommended Validity",
                   "Subject Alternative Names",
                   "Key Usage",
                   "Critical Extensions",
                   "Serial Number",
                   "DB ID (NW)",
                   "DB ID (IP)",
                   "DB ID (HN)",
                   "DB ID (SRV)",
                   "DB ID (CERT)",
                   "Source (NW)",
                   "Source (IP)",
                   "Source (HN)",
                   "Source (SRV)",
                   "Source (CERT)"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                host_names = [mapping.host_name
                              for mapping in host.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                               DnsResourceRecordType.aaaa])]
                network_str = host.ipv4_network.network if host.ipv4_network else None
                network_id = host.ipv4_network.id if host.ipv4_network else None
                network_companies = host.ipv4_network.companies_str if host.ipv4_network else None
                network_sources = host.ipv4_network.sources_str if host.ipv4_network else None
                network_scope = host.ipv4_network.scope_str if host.ipv4_network else None
                host_is_private = host.ip_address.is_private
                host_sources = host.sources_str
                for service in host.services:
                    if service.state in [ServiceState.Open, ServiceState.Closed]:
                        is_http = descriptor.match_nmap_service_name(service)
                        url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"] \
                            if is_http else []
                        for cert_info in service.cert_info:
                            if self._filter(cert_info):
                                matching_host = "n/a"
                                result.append([workspace.name,
                                               "Host",
                                               network_str,
                                               network_scope,
                                               network_companies,
                                               host.address,
                                               host.version_str,
                                               host_is_private,
                                               host.in_scope,
                                               host.os_family,
                                               host.os_details,
                                               None,
                                               None,
                                               host.address,
                                               host.in_scope,
                                               None,
                                               None,
                                               None,
                                               service.protocol_str,
                                               service.port,
                                               service.protocol_port_str,
                                               service.service_name_with_confidence,
                                               service.service_confidence,
                                               service.state_str,
                                               service.nmap_service_state_reason,
                                               service.nmap_product_version,
                                               is_http,
                                               url_str[0] if url_str else None,
                                               matching_host,
                                               cert_info.common_name.lower(),
                                               cert_info.issuer_name.lower(),
                                               cert_info.is_self_signed(),
                                               cert_info.signature_asym_algorithm_str,
                                               cert_info.signature_bits,
                                               cert_info.signature_asym_algorithm_summary,
                                               None,
                                               cert_info.hash_algorithm_str,
                                               cert_info.has_weak_signature(),
                                               cert_info.cert_type_str,
                                               cert_info.valid_from_str,
                                               cert_info.valid_until_str,
                                               cert_info.validity_period_days / 365,
                                               cert_info.is_valid(),
                                               cert_info.has_recommended_duration(),
                                               cert_info.subject_alt_names_str.lower(),
                                               cert_info.key_usage_str,
                                               ", ".join(cert_info.critical_extension_names),
                                               cert_info.serial_number,
                                               network_id,
                                               host.id,
                                               None,
                                               service.id,
                                               cert_info.id,
                                               network_sources,
                                               host_sources,
                                               None,
                                               service.sources_str,
                                               cert_info.sources_str])
                for host_name in host_names:
                    environment = self._domain_config.get_environment(host_name)
                    host_name_sources = host_name.sources_str
                    network_str = host.ipv4_network.network if host.ipv4_network else None
                    for service in host_name.services:
                        if service.state in [ServiceState.Open, ServiceState.Closed] and \
                                descriptor.match_nmap_service_name(service):
                            is_http = descriptor.match_nmap_service_name(service)
                            url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"] \
                                if is_http else []
                            for cert_info in service.cert_info:
                                if self._filter(cert_info):
                                    matching_host = cert_info.matches_host_name(host_name) \
                                        if cert_info.cert_type == CertType.identity else None
                                    result.append([workspace.name,
                                                   "VHost",
                                                   network_str,
                                                   network_scope,
                                                   network_companies,
                                                   host.address,
                                                   host.version_str,
                                                   host_is_private,
                                                   host.in_scope,
                                                   host.os_family,
                                                   host.os_details,
                                                   host_name.domain_name.name,
                                                   host_name.domain_name.scope_str,
                                                   host_name.full_name,
                                                   host_name._in_scope,
                                                   host_name.name,
                                                   host_name.companies_str,
                                                   environment,
                                                   service.protocol_str,
                                                   service.port,
                                                   service.protocol_port_str,
                                                   service.service_name_with_confidence,
                                                   service.service_confidence,
                                                   service.state_str,
                                                   service.nmap_service_state_reason,
                                                   service.nmap_product_version,
                                                   True,
                                                   url_str[0] if url_str else None,
                                                   matching_host,
                                                   cert_info.common_name.lower(),
                                                   cert_info.issuer_name.lower(),
                                                   cert_info.is_self_signed(),
                                                   cert_info.signature_asym_algorithm_str,
                                                   cert_info.signature_bits,
                                                   cert_info.signature_asym_algorithm_summary,
                                                   None,
                                                   cert_info.hash_algorithm_str,
                                                   cert_info.has_weak_signature(),
                                                   cert_info.cert_type_str,
                                                   cert_info.valid_from_str,
                                                   cert_info.valid_until_str,
                                                   cert_info.validity_period_days / 365,
                                                   cert_info.is_valid(),
                                                   cert_info.has_recommended_duration(),
                                                   cert_info.subject_alt_names_str.lower(),
                                                   cert_info.key_usage_str,
                                                   ", ".join(cert_info.critical_extension_names),
                                                   cert_info.serial_number,
                                                   network_id,
                                                   host.id,
                                                   host_name.id,
                                                   service.id,
                                                   cert_info.id,
                                                   network_sources,
                                                   host_sources,
                                                   host_name_sources,
                                                   service.sources_str,
                                                   cert_info.sources_str])
        return result

    def _final_report_host_name_coverage(self, workbook: Workbook):
        result = [["IP Address", "Service", "Host Names", "Common Name", "Subject Alternative\nNames", "Full\nCoverage"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Hostnamen", "Common Name", "Subject Alternative\nNames", "Volle\nAbdeckung"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                host_names = [item.host_name.full_name
                              for item in host.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                            DnsResourceRecordType.aaaa])]
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            matching_host = cert_info.matches_host_names(host_names)
                            result.append([host.address,
                                           service.protocol_port_str,
                                           ", ".join(host_names),
                                           cert_info.common_name,
                                           cert_info.subject_alt_names_str,
                                           self.TRUE if matching_host else None])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Name Coverage",
                                  title="",
                                  description="")

    def _final_report_valid_ca(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Common Name", "Issuer", "Self-\nSigned"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Common Name", "Issuer", "Selbst\nSigniert"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           cert_info.common_name,
                                           cert_info.issuer_name,
                                           self.TRUE if cert_info.is_self_signed() else None])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Valid CAs",
                                  title="",
                                  description="")

    def _final_report_signature_algorithm(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Public Key\nAlgorithm", "Hash Algorithm"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Public-Key-\nAlgorithmus", "Hashalgorithmus"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           cert_info.signature_asym_algorithm_summary,
                                           cert_info.hash_algorithm_str])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Signing Algorithms",
                                  title="",
                                  description="")

    def _final_report_durations(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Valid\nFrom", "Valid\nUntil", "Valid", "Years", "Valid\nDuration"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Gültig\nvon", "Gültig\nbis", "Gültig", "Jahre", "Gültige\nDauer"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name,
                                           cert_info.valid_from_str,
                                           cert_info.valid_until_str,
                                           self.TRUE if cert_info.is_valid() else None,
                                           "{:.2f}".format(cert_info.validity_period_days / 365),
                                           self.TRUE if cert_info.has_recommended_duration() else None])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Durations",
                                  title="",
                                  description="")

    def _final_report_key_usages(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Key Usage"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Verwendungszweck"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           cert_info.key_usage_str])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Key Usage",
                                  title="",
                                  description="")

    def _final_report_critical_extensions(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Critical Extensions"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Kritische Erweiterungen"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           ", ".join(cert_info.critical_extension_names)])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Crit. Extensions",
                                  title="",
                                  description="")

    def final_report(self, workbook: Workbook):
        """
        This method creates all tables that are relevant to the final report.
        """
        self._final_report_host_name_coverage(workbook)
        self._final_report_valid_ca(workbook)
        self._final_report_signature_algorithm(workbook)
        self._final_report_durations(workbook)
        self._final_report_key_usages(workbook)
        self._final_report_critical_extensions(workbook)