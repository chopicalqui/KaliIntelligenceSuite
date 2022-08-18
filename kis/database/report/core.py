# -*- coding: utf-8 -*-
""""This file contains core functionalities for reports."""

__author__ = "Lukas Reiter"
__license__ = "GPL v3.0"
__copyright__ = """Copyright 2022 Lukas Reiter

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""
__version__ = 0.1

import os
import csv
import sys
import re
import enum
import pkgutil
import argparse
import importlib
from database.config import DomainConfig
from database.config import SortingHelpFormatter
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import IllegalCharacterError
from typing import List
from database.model import Host
from database.model import Command
from database.model import FontColor
from database.model import Workspace
from database.model import VhostChoice
from database.model import ServiceState
from database.model import ProtocolType
from database.model import CollectorName
from database.model import ReportScopeType
from database.model import ReportVisibility
from collectors.os.modules.http.core import HttpServiceDescriptor
from sqlalchemy.orm.session import Session


class ReportLanguage(enum.Enum):
    en = enum.auto()
    de = enum.auto()

    def __str__(self):
        return self.name.lower()

    def __repr__(self):
        return str(self)

    @staticmethod
    def argparse(s):
        try:
            return ReportLanguage[s]
        except KeyError:
            return s


class ExcelReport(enum.Enum):
    network = enum.auto()
    host = enum.auto()
    domain = enum.auto()
    hostname = enum.auto()
    service = enum.auto()
    cname = enum.auto()
    email = enum.auto()
    company = enum.auto()
    path = enum.auto()
    credential = enum.auto()
    additionalinfo = enum.auto()
    file = enum.auto()
    breach = enum.auto()
    vulnerability = enum.auto()
    command = enum.auto()
    tls = enum.auto()
    cert = enum.auto()
    leaves = enum.auto()
    dnsrecord = enum.auto()


class BaseReport:
    """
    This class implements all base functionality for generating reports
    """

    TRUE = "•"

    def __init__(self,
                 args,
                 session: Session,
                 name: str,
                 description: str,
                 workspaces: List[Workspace],
                 title: str,
                 **kwargs) -> None:
        self._domain_config = DomainConfig()
        self._name = name
        self._args = args
        self._scope = ReportScopeType[args.scope] if "scope" in args and getattr(args, "scope") else None
        self._visibility = ReportVisibility[args.visibility] \
            if "visibility" in args and getattr(args, "visibility") else None
        protocols = args.protocol if "protocol" in args and getattr(args, "protocol") else []
        self._protocols = [ProtocolType[item] for item in protocols]
        self._session = session
        self._workspaces = workspaces
        self._kwargs = kwargs
        self._not_grep = args.grep_not if "grep_not" in args else False
        self.description = description
        self.title = title
        self._color = "nocolor" in args and not getattr(args, "nocolor")
        if "filter" in args and args.filter:
            self._included_items = [item[1:] for item in args.filter if item[0] == '+']
            self._excluded_items = [item for item in args.filter if item[0] != '+']
        else:
            self._included_items = []
            self._excluded_items = []
        self._included_collectors = args.include if "include" in args and args.include else []
        if "exclude" in args and args.exclude:
            self._excluded_collectors = args.exclude if "all" not in args.exclude else \
                [item.name for item in session.query(CollectorName).all()]
        else:
            self._excluded_collectors = []
        if "grep" in args and args.grep:
            self._regex_list = [re.compile(item) for item in args.grep]
        elif "igrep" in args and args.igrep:
            self._regex_list = [re.compile(item, re.IGNORECASE) for item in args.igrep]
        else:
            self._regex_list = []

    @property
    def scope(self) -> ReportScopeType:
        return self._args.scope if "scope" in self._args else None

    def _egrep(self, results: List[str]) -> List[str]:
        """
        This method returns all lines matching the given list of regular expressions
        :param results: List of strings on which the regular expressions are applied
        :return:
        """
        result = []
        for line in results:
            for regex in self._regex_list:
                positions = []
                for match in regex.finditer(line):
                    if "output" in match.groupdict():
                        result.append(match.group("output"))
                    else:
                        positions.append([match.start(), match.end()])
                if positions:
                    if self._color:
                        position = 0
                        new_line = ""
                        color = FontColor.RED + FontColor.BOLD
                        for start, end in positions:
                            new_line += line[position:start]
                            new_line += color
                            new_line += line[start:end]
                            new_line += FontColor.END
                            position += end
                        new_line += line[position:]
                        line = new_line
                    result.append(line)
        return result

    def fill_excel_sheet(self,
                         worksheet: Worksheet,
                         csv_list: list,
                         name: str = None,
                         title: str = None,
                         description: str = None) -> None:
        """
        This method adds an additional sheet to the given workbook
        :return:
        """
        start_row = 1
        name = name if name is not None else self._name
        title = title if title is not None else self.title
        description = description if description is not None else self.description
        worksheet.title = name
        if description:
            csv_list.insert(0, [])
            csv_list.insert(0, [description])
            start_row += 2
        if title:
            csv_list.insert(0, [])
            csv_list.insert(0, [title])
            start_row += 2
        for row in csv_list:
            try:
                worksheet.append(row)
            except IllegalCharacterError:
                print("ignoring row due to illegal character: {}".format(row), file=sys.stderr)
            except ValueError:
                raise ValueError("cannot add row to sheet '{}': {}".format(self._name, row))
        dimension = worksheet.calculate_dimension()
        dimension = "A{}:{}".format(start_row, dimension.split(":")[-1])
        table = Table(displayName=self._name.replace(" ", ""), ref=dimension)
        style = TableStyleInfo(name="TableStyleLight8")
        table.tableStyleInfo = style
        worksheet.add_table(table)

    def export(self):
        """
        This method executes the export
        :return:
        """
        if "text" in self._args and getattr(self._args, "text"):
            for line in self.get_text():
                print(line)
        elif "grep" in self._args and getattr(self._args, "grep") or \
             "igrep" in self._args and getattr(self._args, "igrep"):
            results = self.grep_text()
            csv_writer = csv.writer(sys.stdout, dialect='excel')
            csv_writer.writerows(results)
        elif "file" in self._args.module and "export_path" in self._args and getattr(self._args, "export_path"):
            if os.path.isdir(self._args.export_path):
                self.export_files()
            else:
                print("Invalid output directory '{}'.".format(self._args.export_path), file=sys.stderr)
        elif "csv" in self._args and getattr(self._args, "csv"):
            results = self.get_csv()
            csv_writer = csv.writer(sys.stdout, dialect='excel')
            csv_writer.writerows(results)

    def _get_unique_file_name(self, output_path: str, file_name: str) -> str:
        """This method returns a unique output path"""
        tmp = os.path.splitext(file_name)
        ext = tmp[-1]
        name = tmp[0]
        path = os.path.join(output_path, "{}{}".format(name, ext))
        i = 1
        while os.path.exists(path):
            new_name = "{}_{:03d}".format(name, i)
            path = os.path.join(output_path, "{}{}".format(new_name, ext))
            i += 1
        return path

    @staticmethod
    def get_add_argparse_arguments(arg_group: argparse.ArgumentParser):
        """
        This method adds the report's specific command line arguments.
        """
        raise NotImplementedError("not implemented")

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text. Method is implemented by subclass.
        :return:
        """
        raise NotImplementedError("not implemented")

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV. Method is implemented by subclass.
        :return:
        """
        raise NotImplementedError("not implemented")

    def grep_text(self) -> List[List[str]]:
        """
        This method returns all information as a list of text.
        :return:
        """
        raise NotImplementedError("not implemented")

    def export_files(self) -> None:
        """
        Exports all files from the database. Method is implemented by subclass.
        :return:
        """
        raise NotImplementedError("not implemented")

    def final_report(self, workbook: Workbook):
        """
        This method creates all tables that are relevant to the final report.
        """
        pass


class ReportClassInfo:
    """This class holds information about existing report modules."""

    def __init__(self, report_class):
        self.report_class = report_class
        self.name = report_class.__module__.split(".")[-1]
        self.description = report_class.__doc__.strip()

    def create_instance(self, **kwargs) -> None:
        """
        This method creates an instance of the current type using the given arguments.
        :param kwargs: The constructor arguments to initialize the class
        :return:
        """
        return self.report_class(**kwargs)


class ReportGenerator:
    """
    This method creates all functionalities to create reports.
    """

    def __init__(self, report_classes: dict):
        self._report_classes = report_classes

    @staticmethod
    def add_argparser_arguments(sub_parser: argparse._SubParsersAction = None) -> dict:
        """
        This method adds all collector arguments to the given ArgumentParser class.
        :param sub_parser: The parser group where the arguments of all collectors shall be added
        :return:
        """
        result = ReportGenerator._load_reports()
        if sub_parser:
            for name, item in result.items():
                report_group = sub_parser.add_parser(name, help=item.description)
                item.report_class.get_add_argparse_arguments(report_group)
        return result

    @staticmethod
    def get_report_argument_parser(epilog: str = None) -> argparse.ArgumentParser:
        """
        This method creates and initializes kisreport's argparser.
        """
        parser = argparse.ArgumentParser(description=__doc__, formatter_class=SortingHelpFormatter, epilog=epilog)
        parser.add_argument("--nocolor", action="store_true", help="disable colored output")
        parser.add_argument("-l", "--list", action='store_true', help="list existing workspaces")
        parser.add_argument('--testing',
                            action="store_true",
                            help="if specified, then KIS uses the testing instead of the production database")
        return parser

    @staticmethod
    def add_sub_argument_parsers(parser: argparse.ArgumentParser) -> argparse._SubParsersAction:
        """
        This method adds kisreport's generic sub_parsers.
        """
        sub_parser = parser.add_subparsers(help='list of available database modules', dest="module")
        # setup excel parser
        parser_excel = sub_parser.add_parser('excel', help='allows writing all identified information into a '
                                                           'microsoft excel file')
        parser_excel.add_argument('FILE', type=str,
                                  help="the path to the microsoft excel file")
        parser_excel.add_argument("-w", "--workspaces",
                                  metavar="WORKSPACE",
                                  help="query the given workspaces",
                                  nargs="+",
                                  type=str)
        parser_excel.add_argument('--filter', metavar='DOMAIN|HOSTNAME|IP|NETWORK|EMAIL', type=str, nargs='*',
                                  help='list of second-level domains (e.g., megacorpone.com), host names '
                                       '(e.g., www.megacorpone.com), IP addresses (e.g., 192.168.1.1), networks (e.g., '
                                       '192.168.0.0/24), or email addresses (e.g., test@megacorpone.com) whose '
                                       'information shall be returned. per default, mentioned items are excluded. add + '
                                       'in front of each item (e.g., +192.168.0.1) to return only these items')
        parser_excel.add_argument('--scope', choices=[item.name for item in ReportScopeType],
                                  help='return only in scope (within) or out of scope (outside) items. per default, '
                                       'all information is returned')
        parser_excel.add_argument('--reports', choices=[item.name for item in ExcelReport],
                                  nargs="+",
                                  default=[item.name for item in ExcelReport if item != ExcelReport.leaves],
                                  help='import only the following reports into Microsoft Excel')
        parser_excel.add_argument("-r", "--report-level",
                                  choices=[item.name for item in VhostChoice],
                                  default=VhostChoice.all.name,
                                  help="specifies the information that shall be displayed in the sheet 'service info'.")
        parser_excel.add_argument('--nosum', action='store_true',
                                  help="there are cases where an in-scope network contains several subnetworks."
                                       "per default, the network report summarizes the statistics into the largest "
                                       "in-scope network and does not show the corresponding subnetworks. if the "
                                       "subnetworks shall be displayed as well, then use this argument")
        parser_excel.add_argument('-p', '--protocol',
                                  choices=[item.name for item in ProtocolType],
                                  default=[item.name for item in ProtocolType],
                                  help="create the service statistics for the following ISO/OSI layer 4 protocols")
        # setup final parser
        parser_final = sub_parser.add_parser('final',
                                             help='allows writing final report tables into microsoft excel file')
        parser_final.add_argument('FILE', type=str,
                                  help="the path to the microsoft excel file")
        parser_final.add_argument("-w", "--workspaces",
                                  metavar="WORKSPACE",
                                  help="query the given workspaces",
                                  nargs="+",
                                  type=str)
        parser_final.add_argument('-l', '--language',
                                  type=ReportLanguage.argparse,
                                  choices=list(ReportLanguage),
                                  default=ReportLanguage.en,
                                  help="the final report's language")
        return sub_parser

    @staticmethod
    def _load_reports() -> dict:
        """
        This method enumerates all report plugins with name ReportClass located in database.report.

        These classes are then used to initialize the command line parser. Based on the user's selection, the desired
        classes are then initialized in method init for data collection.

        :return: A dictionary containing all report classes
        """
        return_value = {}
        module_paths = [""]
        module_paths.extend(os.listdir(os.path.dirname(__file__)))
        for item in module_paths:
            module_path = os.path.join(os.path.dirname(__file__), item)
            if os.path.isdir(module_path) and item != "__pycache__":
                for importer, package_name, _ in pkgutil.iter_modules([module_path]):
                    import_string = "database.report."
                    import_string += "{}.{}".format(item, package_name) if item else package_name
                    module = importlib.import_module(import_string)
                    if "ReportClass" in vars(module):
                        class_ = getattr(module, "ReportClass")
                        report_info = ReportClassInfo(class_)
                        return_value[report_info.name] = report_info
        return return_value

    def create_report_instance(self, args: dict, session: Session, workspaces: List[Workspace]):
        """
        This method can be used by unittest to obtain a fully initialized report instance.
        """
        return self._report_classes[args.module].create_instance(session=session, workspaces=workspaces, args=args)

    def run(self, args: dict, session: Session, workspaces: List[Workspace]) -> None:
        """
        This method initializes the selected report and runs the desired export
        :param args: The argparser arguments based on which the report is created
        :return:
        """
        if args.module == "excel":
            if os.path.exists(args.FILE):
                os.unlink(args.FILE)
            workbook = Workbook()
            first = True
            processed = []
            for report_str in args.reports:
                if report_str not in processed:
                    print("* creating report for: {}".format(report_str))
                    processed.append(report_str)
                    report = self._report_classes[report_str].create_instance(session=session,
                                                                              workspaces=workspaces,
                                                                              args=args)
                    csv_list = report.get_csv()
                    if len(csv_list) > 1:
                        if first:
                            report.fill_excel_sheet(workbook.active, csv_list=csv_list)
                            first = False
                        else:
                            report.fill_excel_sheet(workbook.create_sheet(), csv_list=csv_list)
            workbook.save(args.FILE)
        elif args.module == "final":
            if os.path.exists(args.FILE):
                os.unlink(args.FILE)
            workbook = Workbook()
            workbook.remove(workbook.active)
            for item in ExcelReport:
                report = self._report_classes[item.name].create_instance(session=session,
                                                                         workspaces=workspaces,
                                                                         args=args)
                report.final_report(workbook=workbook)
            workbook.save(args.FILE)
        else:
            report = self._report_classes[args.module].create_instance(session=session,
                                                                       workspaces=workspaces,
                                                                       args=args)
            report.export()


class ServiceStatistics:
    """
    This class maintains all statistics for a network
    """

    def __init__(self, protocols: list):
        self.descriptor = HttpServiceDescriptor()
        self._protocols = protocols
        self.no_open_services = 0
        self.no_open_web_services = 0
        self.no_open_in_scope_services = 0
        self.no_open_in_scope_web_services = 0
        self.no_closed_services = 0
        self.no_closed_web_services = 0
        self.no_closed_in_scope_services = 0
        self.no_closed_in_scope_web_services = 0

    def compute(self, host: Host):
        """
        Compute statistics for the given network.
        """
        # Obtain statistics about services
        for service in host.services:
            if service.protocol in self._protocols:
                in_scope = service.host.in_scope
                is_http = self.descriptor.match_nmap_service_name(service)
                if service.state == ServiceState.Open:
                    self.no_open_services += 1
                    if in_scope:
                        self.no_open_in_scope_services += 1
                    if is_http:
                        self.no_open_web_services += 1
                        if in_scope:
                            self.no_open_in_scope_web_services += 1
                else:
                    self.no_closed_services += 1
                    if in_scope:
                        self.no_closed_in_scope_services += 1
                    if is_http:
                        self.no_closed_web_services += 1
                        if in_scope:
                            self.no_closed_in_scope_web_services += 1
