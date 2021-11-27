# -*- coding: utf-8 -*-
""""This file contains all functionality to convert the data of the database into a JSON object."""

__author__ = "Lukas Reiter"
__license__ = "GPL v3.0"
__copyright__ = """Copyright 2018 Lukas Reiter

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""
__version__ = 0.1

import os
import csv
import sys
import json
import re
import enum
from database.config import DomainConfig
from database.model import FileType
from database.model import CommandFileMapping
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import IllegalCharacterError
from typing import List
from typing import Dict
from database.model import CollectorName
from database.model import Command
from database.model import Host
from database.model import Workspace
from database.model import HostName
from database.model import DomainName
from database.model import Email
from database.model import Company
from database.model import Network
from database.model import Path
from database.model import Credentials
from database.model import AdditionalInfo
from database.model import CollectorType
from database.model import CertInfo
from database.model import TlsInfo
from database.model import ReportScopeType
from database.model import ReportVisibility
from database.model import DnsResourceRecordType
from database.model import Service
from database.model import CertType
from database.model import FontColor
from database.model import PathType
from database.model import ServiceState
from database.model import CipherSuiteSecurity
from collectors.core import BaseUtils
from collectors.os.modules.http.core import HttpServiceDescriptor
from collectors.apis.haveibeenpwned import HaveIBeenPwnedPasteAcccount
from collectors.apis.haveibeenpwned import HaveIBeenPwnedBreachedAcccount
from sqlalchemy.orm.session import Session


class ReportLanguage(enum.Enum):
    en = enum.auto()
    de = enum.auto()

    def __str__(self):
        return self.name.lower()

    def __repr__(self):
        return str(self)

    @staticmethod
    def argparse(s):
        try:
            return ReportLanguage[s]
        except KeyError:
            return s


class ExcelReport(enum.Enum):
    host = enum.auto()
    vhost = enum.auto()
    domain = enum.auto()
    cname = enum.auto()
    network = enum.auto()
    email = enum.auto()
    company = enum.auto()
    path = enum.auto()
    credential = enum.auto()
    additionalinfo = enum.auto()
    file = enum.auto()
    breach = enum.auto()
    vulnerability = enum.auto()
    command = enum.auto()
    tls = enum.auto()
    cert = enum.auto()


class _BaseReportGenerator:
    """
    This class implements all base functionality for generating reports
    """

    TRUE = "•"

    def __init__(self,
                 args,
                 session: Session,
                 workspaces: List[Workspace],
                 name: str,
                 description: str,
                 title: str,
                 **kwargs) -> None:
        self._domain_config = DomainConfig()
        self._name = name
        self._args = args
        self._scope = ReportScopeType[args.scope] if "scope" in args and getattr(args, "scope") else None
        self._visibility = ReportVisibility[args.visibility] \
            if "visibility" in args and getattr(args, "visibility") else None
        self._session = session
        self._workspaces = workspaces
        self._kwargs = kwargs
        self._not_grep = args.grep_not if "grep_not" in args else False
        self.description = description
        self.title = title
        self._color = "nocolor" in args and not getattr(args, "nocolor")
        if "filter" in args and args.filter:
            self._included_items = [item[1:] for item in args.filter if item[0] == '+']
            self._excluded_items = [item for item in args.filter if item[0] != '+']
        else:
            self._included_items = []
            self._excluded_items = []
        self._included_collectors = args.include if "include" in args and args.include else []
        if "exclude" in args and args.exclude:
            self._excluded_collectors = args.exclude if "all" not in args.exclude else \
                [item.name for item in session.query(CollectorName).all()]
        else:
            self._excluded_collectors = []
        if "grep" in args and args.grep:
            self._regex_list = [re.compile(item) for item in args.grep]
        elif "igrep" in args and args.igrep:
            self._regex_list = [re.compile(item, re.IGNORECASE) for item in args.igrep]
        else:
            self._regex_list = []

    @property
    def scope(self) -> ReportScopeType:
        return self._args.scope if "scope" in self._args else None

    def _egrep(self, results: List[str]) -> List[str]:
        """
        This method returns all lines matching the given list of regular expressions
        :param results: List of strings on which the regular expressions are applied
        :return:
        """
        result = []
        for line in results:
            for regex in self._regex_list:
                positions = []
                for match in regex.finditer(line):
                    if "output" in match.groupdict():
                        result.append(match.group("output"))
                    else:
                        positions.append([match.start(), match.end()])
                if positions:
                    if self._color:
                        position = 0
                        new_line = ""
                        color = FontColor.RED + FontColor.BOLD
                        for start, end in positions:
                            new_line += line[position:start]
                            new_line += color
                            new_line += line[start:end]
                            new_line += FontColor.END
                            position += end
                        new_line += line[position:]
                        line = new_line
                    result.append(line)
        return result

    def fill_excel_sheet(self,
                         worksheet: Worksheet,
                         csv_list: list,
                         name: str=None,
                         title: str = None,
                         description: str = None) -> None:
        """
        This method adds an additional sheet to the given workbook
        :return:
        """
        start_row = 1
        name = name if name is not None else self._name
        title = title if title is not None else self.title
        description = description if description is not None else self.description
        worksheet.title = name
        if description:
            csv_list.insert(0, [])
            csv_list.insert(0, [description])
            start_row += 2
        if title:
            csv_list.insert(0, [])
            csv_list.insert(0, [title])
            start_row += 2
        for row in csv_list:
            try:
                worksheet.append(row)
            except IllegalCharacterError:
                print("ignoring row due to illegal character: {}".format(row), file=sys.stderr)
            except ValueError:
                raise ValueError("cannot add row to sheet '{}': {}".format(self._name, row))
        dimension = worksheet.calculate_dimension()
        dimension = "A{}:{}".format(start_row, dimension.split(":")[-1])
        table = Table(displayName=self._name.replace(" ", ""), ref=dimension)
        style = TableStyleInfo(name="TableStyleLight8")
        table.tableStyleInfo = style
        worksheet.add_table(table)

    def export(self):
        """
        This method executes the export
        :return:
        """
        if "csv" in self._args and getattr(self._args, "csv"):
            results = self.get_csv()
            csv_writer = csv.writer(sys.stdout, dialect='excel')
            csv_writer.writerows(results)
        elif "grep" in self._args and getattr(self._args, "grep") or \
             "igrep" in self._args and getattr(self._args, "igrep"):
            results = self.grep_text()
            csv_writer = csv.writer(sys.stdout, dialect='excel')
            csv_writer.writerows(results)
        elif "text" in self._args and getattr(self._args, "text"):
            for line in self.get_text():
                print(line)
        elif "file" in self._args.module:
            if os.path.isdir(self._args.export_path):
                self.export_files()
            else:
                print("Invalid output directory '{}'.".format(self._args.export_path), file=sys.stderr)

    def _get_unique_file_name(self, output_path: str, file_name: str) -> str:
        """This method returns a unique output path"""
        tmp = os.path.splitext(file_name)
        ext = tmp[-1]
        name = tmp[0]
        path = os.path.join(output_path, "{}{}".format(name, ext))
        i = 1
        while os.path.exists(path):
            new_name = "{}_{:03d}".format(name, i)
            path = os.path.join(output_path, "{}{}".format(new_name, ext))
            i += 1
        return path

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text. Method is implemented by subclass.
        :return:
        """
        raise NotImplementedError("not implemented")

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV. Method is implemented by subclass.
        :return:
        """
        raise NotImplementedError("not implemented")

    def grep_text(self) -> List[List[str]]:
        """
        This method returns all information as a list of text.
        :return:
        """
        raise NotImplementedError("not implemented")

    def export_files(self) -> None:
        """
        Exports all files from the database. Method is implemented by subclass.
        :return:
        """
        raise NotImplementedError("not implemented")

    def final_report(self, workbook: Workbook):
        """
        This method creates all tables that are relevant to the final report.
        """
        pass


class _HostReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for hosts
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="host info",
                         title="IP address details",
                         description="This table provides a consolidated view about all IP addresses (see column "
                                     "'IP Address (IP)'), their virtual hosts (see column 'Host Name (HN)') as well as "
                                     "their identified TCP/UDP services (see columns 'UDP/TCP' and 'Port'). "
                                     ""
                                     "If you are just interested in service information on the IP address level, then "
                                     "filter for value 'Host' in column 'Type'. If you are interested in service "
                                     "information on the virtual host level, filter for value 'VHost' in column "
                                     "'Type'. "
                                     ""
                                     "Note that host ames, which do not resolve to an IP address, are not listed in "
                                     "this sheet; use sheet 'host name info' to analyse them.",
                         **kwargs)

    def _filter(self, host: Host) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        rvalue = host.is_processable(included_items=self._included_items,
                                     excluded_items=self._excluded_items,
                                     scope=self._scope,
                                     include_host_names=True)
        return rvalue

    def _egrep_text(self, host: HostName) -> List[str]:
        """
        This method returns all lines matching the given list of regular expressions
        :param domain: The domain name whose text output shall be parsed
        :return:
        """
        result = self._egrep(host.get_text(ident=0,
                                           exclude_collectors=self._excluded_collectors,
                                           include_collectors=self._included_collectors,
                                           scope=self._scope,
                                           show_metadata=False))
        return result

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        for workspace in self._workspaces:
            for host in workspace.hosts:
                if self._filter(host):
                    rvalue.extend(host.get_text(ident=0,
                                                scope=self._scope,
                                                exclude_collectors=self._excluded_collectors,
                                                include_collectors=self._included_collectors,
                                                report_visibility=self._visibility,
                                                color=self._color))
        return rvalue

    def grep_text(self) -> List[List[str]]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rows = [["DB ID",
                 "Workspace",
                 "Network",
                 "Companies",
                 "In Scope",
                 "Address",
                 "Result"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                if self._filter(host):
                    ipv4_network = None
                    companies = None
                    if host.ipv4_network:
                        ipv4_network = host.ipv4_network.network
                        companies = host.ipv4_network.companies_str
                    results = self._egrep_text(host)
                    if self._not_grep and not results:
                        rows.append([host.id,
                                     workspace.name,
                                     ipv4_network,
                                     companies,
                                     host.in_scope,
                                     host.ip,
                                     None])
                    elif not self._not_grep:
                        for result in results:
                            rows.append([host.id,
                                         workspace.name,
                                         ipv4_network,
                                         companies,
                                         host.in_scope,
                                         host.ip,
                                         result])
        return rows

    def get_csv(self) -> List[List[str]]:
        """
        Method determines whether the given item shall be included into the report
        """
        descriptor = HttpServiceDescriptor()
        result = [["Workspace",
                   "Type",
                   "Network (NW)",
                   "Scope (NW)",
                   "Company (NW)",
                   "IP Address (IP)",
                   "Version (IP)",
                   "Private IP",
                   "In Scope (IP)",
                   "OS Family",
                   "OS Details",
                   "Host Names/IP Addresses",
                   "Second-Level Domain (SLD)",
                   "Scope (SLD)",
                   "Host Name (HN)",
                   "In Scope (HN)",
                   "In Scope (IP or HN)",
                   "Name Only (HN)",
                   "Company (HN)",
                   "Environment (HN)",
                   "UDP/TCP",
                   "Port",
                   "Service (SRV)",
                   "Nmap Name (SRV)",
                   "Confidence (SRV)",
                   "State (SRV)",
                   "Reason State",
                   "Banner Information",
                   "TLS",
                   "Is HTTP",
                   "URL",
                   "SMB Message Signing",
                   "RD NLA",
                   "DB ID (NW)",
                   "DB ID (IP)",
                   "DB ID (HN)",
                   "DB ID (SRV)",
                   "Source (NW)",
                   "Source (IP)",
                   "Source (HN)",
                   "Source (SRV)",
                   "No. Commands",
                   "No. Vulnerabilities"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                if self._filter(host):
                    host_names = [mapping.host_name
                                  for mapping in host.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                                   DnsResourceRecordType.aaaa])]
                    host_names_str = ", ".join([item.full_name for item in host_names])
                    network_str = host.ipv4_network.network if host.ipv4_network else None
                    network_id = host.ipv4_network.id if host.ipv4_network else None
                    network_companies = host.ipv4_network.companies_str if host.ipv4_network else None
                    network_sources = host.ipv4_network.sources_str if host.ipv4_network else None
                    network_scope = host.ipv4_network.scope_str if host.ipv4_network else None
                    host_is_private = host.ip_address.is_private
                    host_sources = host.sources_str
                    services_exist = False
                    for service in host.services:
                        if service.state in [ServiceState.Open, ServiceState.Closed]:
                            services_exist = True
                            is_http = descriptor.match_nmap_service_name(service)
                            url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"] \
                                if is_http else []
                            result.append([workspace.name,
                                           "Host",
                                           network_str,
                                           network_scope,
                                           network_companies,
                                           host.address,
                                           host.version_str,
                                           host_is_private,
                                           host.in_scope,
                                           host.os_family,
                                           host.os_details,
                                           host_names_str,
                                           None,
                                           None,
                                           host.address,
                                           host.in_scope,
                                           host.in_scope,
                                           None,
                                           None,
                                           None,
                                           service.protocol_str,
                                           service.port,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           service.service_confidence,
                                           service.state_str,
                                           service.nmap_service_state_reason,
                                           service.nmap_product_version,
                                           service.tls,
                                           is_http,
                                           url_str[0] if url_str else None,
                                           service.smb_message_signing,
                                           service.rdp_nla,
                                           network_id,
                                           host.id,
                                           None,
                                           service.id,
                                           network_sources,
                                           host_sources,
                                           None,
                                           service.sources_str,
                                           len(service.get_completed_commands()),
                                           len(service.vulnerabilities)])
                    for host_name in host_names:
                        environment = self._domain_config.get_environment(host_name)
                        hosts = [mapping.host
                                 for mapping in host_name.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                                       DnsResourceRecordType.aaaa])]
                        hosts_str = ", ".join([item.address for item in hosts])
                        host_name_sources = host_name.sources_str
                        network_str = host.ipv4_network.network if host.ipv4_network else None
                        for service in host_name.services:
                            if service.state in [ServiceState.Open, ServiceState.Closed]:
                                services_exist = True
                                is_http = descriptor.match_nmap_service_name(service)
                                url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"] \
                                    if is_http else []
                                result.append([workspace.name,
                                               "VHost",
                                               network_str,
                                               network_scope,
                                               network_companies,
                                               host.address,
                                               host.version_str,
                                               host_is_private,
                                               host.in_scope,
                                               host.os_family,
                                               host.os_details,
                                               hosts_str,
                                               host_name.domain_name.name,
                                               host_name.domain_name.scope_str,
                                               host_name.full_name,
                                               host_name._in_scope,
                                               host.in_scope or host_name._in_scope,
                                               host_name.name,
                                               host_name.companies_str,
                                               environment,
                                               service.protocol_str,
                                               service.port,
                                               service.protocol_port_str,
                                               service.service_name_with_confidence,
                                               service.service_confidence,
                                               service.state_str,
                                               service.nmap_service_state_reason,
                                               service.nmap_product_version,
                                               service.tls,
                                               is_http,
                                               url_str[0] if url_str else None,
                                               service.smb_message_signing,
                                               service.rdp_nla,
                                               network_id,
                                               host.id,
                                               host_name.id,
                                               service.id,
                                               network_sources,
                                               host_sources,
                                               host_name_sources,
                                               service.sources_str,
                                               len(service.get_completed_commands()),
                                               len(service.vulnerabilities)])
                    if not services_exist:
                        result.append([workspace.name,
                                       "Host",
                                       network_str,
                                       network_scope,
                                       network_companies,
                                       host.address,
                                       host.version_str,
                                       host_is_private,
                                       host.in_scope,
                                       host.os_family,
                                       host.os_details,
                                       host_names_str,
                                       None,
                                       None,
                                       host.address,
                                       host.in_scope,
                                       host.in_scope,
                                       None,
                                       None,
                                       None,
                                       None,
                                       None,
                                       None,
                                       None,
                                       None,
                                       "not scanned",
                                       None,
                                       None,
                                       None,
                                       None,
                                       None,
                                       None,
                                       None,
                                       network_id,
                                       host.id,
                                       None,
                                       None,
                                       network_sources,
                                       host_sources,
                                       None,
                                       None,
                                       0,
                                       0])
                        for host_name in host_names:
                            environment = self._domain_config.get_environment(host_name)
                            host_name_sources = host_name.sources_str
                            hosts = [mapping.host
                                     for mapping in host_name.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                                           DnsResourceRecordType.aaaa])]
                            hosts_str = ", ".join([item.address for item in hosts])
                            result.append([workspace.name,
                                           "VHost",
                                           network_str,
                                           network_scope,
                                           network_companies,
                                           host.address,
                                           host.version_str,
                                           host_is_private,
                                           host.in_scope,
                                           host.os_family,
                                           host.os_details,
                                           hosts_str,
                                           host_name.domain_name.name,
                                           host_name.domain_name.scope_str,
                                           host_name.full_name,
                                           host_name._in_scope,
                                           host.in_scope or host_name._in_scope,
                                           host_name.name,
                                           host_name.companies_str,
                                           environment,
                                           None,
                                           None,
                                           None,
                                           None,
                                           None,
                                           "not scanned",
                                           None,
                                           None,
                                           None,
                                           None,
                                           None,
                                           None,
                                           None,
                                           network_id,
                                           host.id,
                                           host_name.id,
                                           None,
                                           network_sources,
                                           host_sources,
                                           host_name_sources,
                                           None,
                                           0,
                                           0])
        return result

    def final_report(self, workbook: Workbook):
        """
        This method creates all tables that are relevant to the final report.
        """
        descriptor = HttpServiceDescriptor()
        result = [["Type",
                   "IP Address (IP)",
                   "Host Name (HN)",
                   "Service",
                   "Service\nName",
                   "State",
                   "Is\nHTTP",
                   "TLS",
                   "Banner Information"]]
        if self._args.language == ReportLanguage.de:
            result = [["Typ",
                       "IP-Addresse (IP)",
                       "Hostname (HN)",
                       "Service",
                       "Service\nName",
                       "Status",
                       "Ist\nHTTP",
                       "TLS",
                       "Banner-Information"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                if host.in_scope:
                    host_names = [mapping.host_name
                                  for mapping in host.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                                   DnsResourceRecordType.aaaa])]
                    for service in host.services:
                        if service.state in [ServiceState.Open, ServiceState.Closed]:
                            is_http = descriptor.match_nmap_service_name(service)
                            result.append(["Host",
                                           host.address,
                                           host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           service.state_str,
                                           self.TRUE if is_http else None,
                                           self.TRUE if service.tls else None,
                                           service.nmap_product_version])
                    for host_name in host_names:
                        for service in host_name.services:
                            if service.state in [ServiceState.Open, ServiceState.Closed] and \
                                    descriptor.match_nmap_service_name(service):
                                is_http = descriptor.match_nmap_service_name(service)
                                result.append(["VHost",
                                               host.address,
                                               host_name.full_name,
                                               service.protocol_port_str,
                                               service.service_name_with_confidence,
                                               service.state_str,
                                               self.TRUE if is_http else None,
                                               self.TRUE if service.tls else None,
                                               service.nmap_product_version])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Service Results",
                                  title="",
                                  description="")


class _HostNameReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for host names
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="host name info",
                         title="Host name details",
                         description="The table provides details about all identified host names (see column "
                                     "'Host Name (HN)') and their respective services (see columns 'UDP/TCP' and "
                                     "'Port'). "
                                     ""
                                     "Note that column 'In Scope (HN)' is true, if the respective host name itself "
                                     "is in scope of the engagement. Column 'In Scope (VHost)' is true, if column "
                                     "'In Scope (HN)' is true and the respective host name resolves to at least one "
                                     "IP address, which is in scope of the engagement as well.",
                         **kwargs)

    def _filter(self, host_name: HostName) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return host_name.is_processable(included_items=self._included_items,
                                        excluded_items=self._excluded_items,
                                        collector_type=CollectorType.vhost_service,
                                        scope=self._scope,
                                        include_ip_address=True)

    def _egrep_text(self, host_name: HostName) -> List[str]:
        """
        This method returns all lines matching the given list of regular expressions
        :param host_name: The host_name name whose text output shall be parsed
        :return:
        """
        result = self._egrep(host_name.get_text(ident=0,
                                                exclude_collectors=self._excluded_collectors,
                                                include_collectors=self._included_collectors,
                                                scope=self._scope,
                                                show_metadata=False))
        return result

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        companies = {}
        for item in self._session.query(HostName).filter(HostName.name.is_(None)).all():
            companies[item.domain_name.name] = item.companies_str
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                for item in domain.host_names:
                    if self._filter(item):
                        rvalue.extend(item.get_text(ident=0,
                                                    scope=self._scope,
                                                    companies=companies,
                                                    exclude_collectors=self._excluded_collectors,
                                                    include_collectors=self._included_collectors,
                                                    report_visibility=self._visibility,
                                                    color=self._color))
        return rvalue

    def grep_text(self) -> List[List[str]]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rows = [["DB ID",
                 "Workspace",
                 "Domain Name",
                 "Companies",
                 "In Scope",
                 "Host Name",
                 "Host Name Summary",
                 "Result"]]
        companies = {}
        for item in self._session.query(HostName).filter(HostName.name.is_(None)).all():
            companies[item.domain_name.name] = item.companies_str
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    if self._filter(host_name):
                        in_scope = host_name.in_scope(CollectorType.vhost_service)
                        results = self._egrep_text(host_name)
                        if self._not_grep and not results:
                            rows.append([host_name.id,
                                         workspace.name,
                                         domain.name,
                                         companies[host_name.domain_name.name],
                                         in_scope,
                                         host_name.full_name,
                                         host_name.summary,
                                         None])
                        elif not self._not_grep:
                            for row in results:
                                rows.append([host_name.id,
                                             workspace.name,
                                             domain.name,
                                             companies[host_name.domain_name.name],
                                             in_scope,
                                             host_name.full_name,
                                             host_name.summary,
                                             row])
        return rows

    def get_csv(self) -> List[List[str]]:
        """
        Method determines whether the given item shall be included into the report
        """
        rvalue = [["Workspace",
                   "Second-Level Domain (SLD)",
                   "Scope (SLD)",
                   "Companies (SLD)",
                   "Host Name (HN)",
                   "In Scope (HN)",
                   "In Scope (Vhost)",
                   "Name Only (HN)",
                   "Environment (HN)",
                   "Summary (HN)",
                   "IP Addresses",
                   "Service Summary",
                   "TCP/UDP",
                   "Port",
                   "Service (SRV)",
                   "State (SRV)",
                   "Sources (SRV)",
                   "TLS",
                   "Reason",
                   "Name (SRV)",
                   "Confidence (SRV)",
                   "NMap Service Name Original",
                   "Nmap Product",
                   "Version",
                   "Product Summary",
                   "SMB Message Signing",
                   "RDP NLA",
                   "Sources (Host Name)",
                   "Sources (Service)",
                   "No. Commands (Service)",
                   "No. Vulnerabilities (Service)",
                   "DB ID (SLD)",
                   "DB ID (HN)"]]
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    if host_name.name and self._filter(host_name):
                        ipv4_addresses = host_name.get_host_host_name_mappings_str(types=[DnsResourceRecordType.a,
                                                                                          DnsResourceRecordType.aaaa])
                        host_name_sources = host_name.sources_str
                        in_scope = host_name.in_scope(CollectorType.vhost_service)
                        environment = self._domain_config.get_environment(host_name)
                        if host_name.services:
                            for service in host_name.services:
                                rvalue.append([workspace.name,
                                               domain.name,
                                               domain.scope_str,
                                               domain.companies_str,
                                               host_name.full_name,
                                               host_name._in_scope,
                                               in_scope,
                                               host_name.name,
                                               environment,
                                               host_name.summary,
                                               ipv4_addresses,
                                               service.summary,
                                               service.protocol_str,
                                               service.port,
                                               service.protocol_port_str,
                                               service.state_str,
                                               service.sources_str,
                                               service.nmap_tunnel,
                                               service.nmap_service_state_reason,
                                               service.service_name_with_confidence,
                                               service.service_confidence,
                                               service.nmap_service_name_original_with_confidence,
                                               service.nmap_product,
                                               service.nmap_version,
                                               service.nmap_product_version,
                                               service.smb_message_signing,
                                               service.rdp_nla,
                                               host_name_sources,
                                               service.sources_str,
                                               len(service.get_completed_commands()),
                                               len(service.vulnerabilities),
                                               domain.id,
                                               host_name.id])
                        else:
                            rvalue.append([workspace.name,
                                           domain.name,
                                           domain.scope_str,
                                           domain.companies_str,
                                           host_name.full_name,
                                           host_name._in_scope,
                                           in_scope,
                                           host_name.name,
                                           environment,
                                           host_name.summary,
                                           ipv4_addresses, None, None, None, None, None, None, None, None, None, None,
                                           None, None, None, None, None, None, host_name_sources, None, None, None,
                                           domain.id,
                                           host_name.id])
            return rvalue


class _DomainNameReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for domains
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="domain info",
                         title="Domain name details",
                         description="The table provides an overview about all identified second-level domains "
                                     "(see column 'Second-Level Domain (SLD)') and their sub-domains (see "
                                     "column 'Host Name (HN)'). In addition, it documents to which IP addresses the "
                                     "respective sub-domains resolve. If no IP address is available, then either "
                                     "collector dnshostpublic or dnshost have not been executed or the sub-domain "
                                     "could not be resolved. "
                                     ""
                                     "Note that column 'In Scope (HN)' is true, if the respective host name itself "
                                     "is in scope of the engagement. Column 'In Scope (VHost)' is true, if column "
                                     "'In Scope (HN)' is true and the respective host name resolves to at least one "
                                     "IP address, which is in scope of the engagement as well.",
                         **kwargs)

    def _filter(self, domain_name: DomainName) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return domain_name.is_processable(included_items=self._included_items,
                                          excluded_items=self._excluded_items,
                                          scope=self._scope,
                                          include_ip_address=True)

    def _egrep_text(self, domain: DomainName) -> List[str]:
        """
        This method returns all lines matching the given list of regular expressions
        :param domain: The domain name whose text output shall be parsed
        :return:
        """
        result = self._egrep(domain.get_text(ident=0,
                                             exclude_collectors=self._excluded_collectors,
                                             include_collectors=self._included_collectors,
                                             scope=self._scope,
                                             show_metadata=False))
        return result

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                if self._filter(domain):
                    rvalue.extend(domain.get_text(ident=0,
                                                  exclude_collectors=self._excluded_collectors,
                                                  include_collectors=self._included_collectors,
                                                  scope=self._scope,
                                                  report_visibility=self._visibility,
                                                  color=self._color))
                    rvalue.append("")
        return rvalue

    def grep_text(self) -> List[List[str]]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rows = [["DB ID (Domain)",
                 "DB ID (Hostname)",
                 "Second-Level Domain",
                 "Second-Level Domain Scope",
                 "Companies",
                 "Result"]]
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                if self._filter(domain):
                    results = self._egrep_text(domain)
                    if self._not_grep and not results:
                        rows.append([domain.id,
                                     workspace.name,
                                     domain.name,
                                     domain.scope_str,
                                     domain.companies_str,
                                     None])
                    elif not self._not_grep:
                        for result in results:
                            rows.append([domain.id,
                                         workspace.name,
                                         domain.name,
                                         domain.scope_str,
                                         domain.companies_str,
                                         result])
        return rows

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["Workspace",
                 "Second-Level Domain (SLD)",
                 "Scope (SLD)",
                 "Host Name (HN)",
                 "In Scope (HN)",
                 "In Scope (Vhost)",
                 "Companies (SLD)",
                 "Sources (HN)",
                 "Name Only (HN)",
                 "Environment (HN)",
                 "Record Type",
                 "Resolves To",
                 "Resolves To In Scope",
                 "Resolves to Network",
                 "Resolves to Companies",
                 "Number of Open Services",
                 "Number of Closed Services",
                 "DB ID (SLD)",
                 "DB ID (HN)"]]
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                if self._filter(domain):
                    for host_name in domain.host_names:
                        sources = host_name.sources_str
                        environment = self._domain_config.get_environment(host_name)
                        printed = False
                        for mapping in host_name.resolved_host_name_mappings:
                            printed = True
                            rows.append([workspace.name,
                                         domain.name,
                                         domain.scope_str,
                                         host_name.full_name,
                                         host_name._in_scope,
                                         host_name.in_scope(CollectorType.vhost_service),
                                         domain.companies_str,
                                         sources,
                                         host_name.name,
                                         environment,
                                         mapping.type_str,
                                         mapping.resolved_host_name.full_name,
                                         mapping.resolved_host_name.in_scope(CollectorType.vhost_service),
                                         None,
                                         mapping.resolved_host_name.domain_name.companies_str
                                         if mapping.resolved_host_name.domain_name else None,
                                         None,
                                         None,
                                         domain.id,
                                         host_name.id])
                        for mapping in host_name.host_host_name_mappings:
                            printed = True
                            open_services = len([service for service in mapping.host.services
                                                 if service.state == ServiceState.Open])
                            closed_services = len([service for service in mapping.host.services
                                                   if service.state == ServiceState.Closed])
                            rows.append([workspace.name,
                                         domain.name,
                                         domain.scope_str,
                                         host_name.full_name,
                                         host_name._in_scope,
                                         host_name.in_scope(CollectorType.vhost_service),
                                         domain.companies_str,
                                         sources,
                                         host_name.name,
                                         environment,
                                         mapping.type_str,
                                         mapping.host.address,
                                         mapping.host.in_scope,
                                         mapping.host.ipv4_network.network if mapping.host.ipv4_network else None,
                                         mapping.host.ipv4_network.companies_str
                                         if mapping.host.ipv4_network else None,
                                         open_services,
                                         closed_services,
                                         domain.id,
                                         host_name.id])
                        if not printed:
                            rows.append([workspace.name,
                                         domain.name,
                                         domain.scope_str,
                                         host_name.full_name,
                                         host_name._in_scope,
                                         host_name.in_scope(CollectorType.vhost_service),
                                         domain.companies_str,
                                         sources,
                                         host_name.name,
                                         environment,
                                         None,
                                         None,
                                         None,
                                         None,
                                         None,
                                         None,
                                         None,
                                         domain.id,
                                         host_name.id])
        return rows

    def final_report(self, workbook: Workbook):
        """
        This method creates all tables that are relevant to the final report.
        """
        result = [["Second-Level Domain",
                   "Host Name",
                   "Record\nType",
                   "Resolves To",
                   "Source\nHost Name"]]
        if self._args.language == ReportLanguage.de:
            result = [["Second-Level Domain",
                       "Hostname",
                       "DNS-\nTyp",
                       "Löst auf zu",
                       "Quelle\nHostname"]]
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    if host_name._in_scope:
                        full_name = host_name.full_name
                        printed = False
                        for mapping in host_name.resolved_host_name_mappings:
                            printed = True
                            result.append([domain.name,
                                           full_name,
                                           mapping.type_str,
                                           mapping.resolved_host_name.full_name,
                                           host_name.sources_str])
                        for mapping in host_name.host_host_name_mappings:
                            printed = True
                            result.append([domain.name,
                                           full_name,
                                           mapping.type_str,
                                           mapping.host.address,
                                           host_name.sources_str])
                        if not printed:
                            result.append([domain.name,
                                           full_name,
                                           None,
                                           None,
                                           host_name.sources_str])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Domain Results",
                                  title="",
                                  description="")


class _CanonicalNameReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for domain name service resolution
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="DNS cname records",
                         title="Overview Canonical Name Records",
                         description="This table summarizes how host names are resolved. Use column 'Resolves to "
                                     "IP' to identify host names that do not resolve to an IP address.",
                         **kwargs)

    def _filter(self, domain_name: DomainName) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return domain_name.is_processable(included_items=self._included_items,
                                          excluded_items=self._excluded_items,
                                          scope=self._scope,
                                          include_ip_address=True)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID (Domain)",
                 "DB ID (Hostname)",
                 "Workspace",
                 "Second-Level Domain",
                 "Second-Level Domain Scope",
                 "Host Name",
                 "Host Name In Scope",
                 "Vhost In-Scope",
                 "CNAME Records",
                 "Resolves to IPv4",
                 "Resolves to IPv6",
                 "Resolves to IP"]]
        for workspace in self._workspaces:
            for host_name in self._session.query(HostName) \
                    .join(DomainName) \
                    .join(Workspace) \
                    .filter(Workspace.name == workspace.name).all():
                if self._filter(host_name.domain_name) and not host_name.source_host_name_mappings:
                    resolves_to_ipv4 = False
                    resolves_to_ipv6 = False
                    cnames = host_name.canonical_name_records
                    if len(cnames) > 1:
                        for item in cnames[-1].host_host_name_mappings:
                            resolves_to_ipv4 |= item.resolves_to_ipv4_address()
                            resolves_to_ipv6 |= item.resolves_to_ipv6_address()
                        rows.append([host_name.domain_name.id,
                                     host_name.id,
                                     workspace.name,
                                     host_name.domain_name.name,
                                     host_name.domain_name.scope_str,
                                     host_name.full_name,
                                     host_name._in_scope,
                                     host_name.in_scope(CollectorType.vhost_service),
                                     " -> ".join([item.summary for item in cnames]),
                                     resolves_to_ipv4,
                                     resolves_to_ipv6,
                                     resolves_to_ipv6 or resolves_to_ipv4])
        return rows


class _PathReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for paths
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="path info",
                         title="Overview Identified Paths",
                         description="The table provides an overview of all identified URLs and network shares. You "
                                     "can use column 'Path Type' to filter for specific paths (e.g., URLs).",
                         **kwargs)
        self._path_types = []
        if "type" in args and args.type:
            self._path_types = [PathType[item] for item in args.type]

    def _filter(self, path: Path) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return (not self._path_types or path.type in self._path_types) and \
               path.is_processable(included_items=self._included_items,
                                   excluded_items=self._excluded_items,
                                   scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["Workspace",
                 "Type",
                 "Network/Second-Level-Domain (NSLD)",
                 "Scope (NSLD)",
                 "Company (NSLD)",
                 "Address",
                 "In Scope (Address)",
                 "In Scope (Vhost)",
                 "Resolves To (RT)",
                 "Summary (IP/Vhost)",
                 "Summary (Service)",
                 "TCP/UDP",
                 "Port",
                 "Type (Path)",
                 "Access Code",
                 "Size Bytes",
                 "Full Path",
                 "Root Directory",
                 "Query",
                 "Sources (Path)",
                 "DB ID (Path)"]]
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    hosts_str = host_name.get_host_host_name_mappings_str([DnsResourceRecordType.a,
                                                                           DnsResourceRecordType.aaaa])
                    for service in host_name.services:
                        for path in service.paths:
                            if self._filter(path):
                                if path.queries:
                                    for query in path.queries:
                                        rows.append([workspace.name,
                                                     "vhost",
                                                     host_name.domain_name.name,
                                                     host_name.domain_name.scope_str,
                                                     host_name.domain_name.companies_str,
                                                     host_name.full_name,
                                                     host_name._in_scope,
                                                     host_name.in_scope(CollectorType.vhost_service),
                                                     hosts_str,
                                                     host_name.summary,
                                                     service.summary,
                                                     service.protocol_str,
                                                     service.port,
                                                     path.type_str,
                                                     path.return_code,
                                                     path.size_bytes,
                                                     path.get_path(),
                                                     path.name == "/",
                                                     query.query,
                                                     path.sources_str,
                                                     path.id])
                                else:
                                    rows.append([workspace.name,
                                                 "vhost",
                                                 host_name.domain_name.name,
                                                 host_name.domain_name.scope_str,
                                                 host_name.domain_name.companies_str,
                                                 host_name.full_name,
                                                 host_name._in_scope,
                                                 host_name.in_scope(CollectorType.vhost_service),
                                                 hosts_str,
                                                 host_name.summary,
                                                 service.summary,
                                                 service.protocol_str,
                                                 service.port,
                                                 path.type_str,
                                                 path.return_code,
                                                 path.size_bytes,
                                                 path.get_path(),
                                                 path.name == "/",
                                                 None,
                                                 path.sources_str,
                                                 path.id])
            for host in workspace.hosts:
                host_names = host.get_host_host_name_mappings_str([DnsResourceRecordType.a,
                                                                   DnsResourceRecordType.aaaa])
                if host.ipv4_network:
                    ipv4_network = host.ipv4_network.network
                    scope = host.ipv4_network.scope_str
                    companies = host.ipv4_network.companies_str
                else:
                    ipv4_network = None
                    scope = None
                    companies = None
                for service in host.services:
                    for path in service.paths:
                        if self._filter(path):
                            if path.queries:
                                for query in path.queries:
                                    rows.append([workspace.name,
                                                 "host",
                                                 ipv4_network,
                                                 scope,
                                                 companies,
                                                 host.address,
                                                 host.in_scope,
                                                 None,
                                                 host_names,
                                                 host.summary,
                                                 service.summary,
                                                 service.protocol_str,
                                                 service.port,
                                                 path.type_str,
                                                 path.return_code,
                                                 path.size_bytes,
                                                 path.get_path(),
                                                 path.name == "/",
                                                 query.query,
                                                 path.sources_str,
                                                 path.id])
                            else:
                                rows.append([workspace.name,
                                             "host",
                                             ipv4_network,
                                             scope,
                                             companies,
                                             host.address,
                                             host.in_scope,
                                             None,
                                             host_names,
                                             host.summary,
                                             service.summary,
                                             service.protocol_str,
                                             service.port,
                                             path.type_str,
                                             path.return_code,
                                             path.size_bytes,
                                             path.get_path(),
                                             path.name == "/",
                                             None,
                                             path.sources_str,
                                             path.id])
        return rows


class _TlsInfoReportGenerator(_BaseReportGenerator):
    """
    This class creates all reports for TLS information
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="tls info",
                         title="Overview TLS Ciphers",
                         description="The table provides an overview of all identified TLS cipher suites.",
                         **kwargs)

    def _filter(self, tls_info: TlsInfo) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return tls_info.is_processable(included_items=self._included_items,
                                       excluded_items=self._excluded_items,
                                       scope=self.scope)

    def _kex_summary(self, kex_algorithm_str: str, kex_algorithm_bits: int) -> str:
        result = None
        if kex_algorithm_str:
            result = kex_algorithm_str
        if kex_algorithm_bits:
            result = result if result else ""
            result += " {}".format(kex_algorithm_bits)
        return result

    def get_csv(self) -> List[List[str]]:
        """
        This method returns TLS information
        :return: List of strings containing TLS properties
        """
        descriptor = HttpServiceDescriptor()
        result = [["Workspace",
                   "Type",
                   "Network (NW)",
                   "Scope (NW)",
                   "Company (NW)",
                   "IP Address (IP)",
                   "Version (IP)",
                   "Private IP",
                   "In Scope (IP)",
                   "OS Family",
                   "OS Details",
                   "Second-Level Domain (SLD)",
                   "Scope (SLD)",
                   "Host Name (HN)",
                   "In Scope (HN)",
                   "Name (HN)",
                   "Company (HN)",
                   "Environment",
                   "UDP/TCP",
                   "Port",
                   "Service (SRV)",
                   "Nmap Name (SRV)",
                   "Confidence (SRV)",
                   "State (SRV)",
                   "Reason State",
                   "Banner Information",
                   "Is HTTP",
                   "URL",
                   "TLS Version",
                   "Preference",
                   "Heartbleed",
                   "Compressors",
                   "Order",
                   "Cipher Suite (IANA)",
                   "Prefered",
                   "KEX Algorithm",
                   "Security",
                   "DB ID (NW)",
                   "DB ID (IP)",
                   "DB ID (HN)",
                   "DB ID (SRV)",
                   "DB ID (TLS)",
                   "Source (NW)",
                   "Source (IP)",
                   "Source (HN)",
                   "Source (SRV)",
                   "Source (TLS)"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                host_names = [mapping.host_name
                              for mapping in host.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                               DnsResourceRecordType.aaaa])]
                network_str = host.ipv4_network.network if host.ipv4_network else None
                network_id = host.ipv4_network.id if host.ipv4_network else None
                network_companies = host.ipv4_network.companies_str if host.ipv4_network else None
                network_sources = host.ipv4_network.sources_str if host.ipv4_network else None
                network_scope = host.ipv4_network.scope_str if host.ipv4_network else None
                host_is_private = host.ip_address.is_private
                host_sources = host.sources_str
                for service in host.services:
                    if service.state in [ServiceState.Open, ServiceState.Closed]:
                        is_http = descriptor.match_nmap_service_name(service)
                        url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"] \
                            if is_http else []
                        patch_missing_kex = {}
                        for tls_info in service.tls_info:
                            if self._filter(tls_info):
                                for mapping in tls_info.cipher_suite_mappings:
                                    cipher_suite = mapping.cipher_suite
                                    key = "{}{}{}{}".format(host.address,
                                                            service.protocol_port_str,
                                                            tls_info.version_str,
                                                            cipher_suite.iana_name)
                                    kex_algorithm = mapping.kex_algorithm_details_str
                                    if kex_algorithm:
                                        patch_missing_kex[key] = kex_algorithm
                                    elif key in patch_missing_kex:
                                        kex_algorithm = patch_missing_kex[key]
                                    result.append([workspace.name,
                                                   "Host",
                                                   network_str,
                                                   network_scope,
                                                   network_companies,
                                                   host.address,
                                                   host.version_str,
                                                   host_is_private,
                                                   host.in_scope,
                                                   host.os_family,
                                                   host.os_details,
                                                   None,
                                                   None,
                                                   host.address,
                                                   host.in_scope,
                                                   None,
                                                   None,
                                                   None,
                                                   service.protocol_str,
                                                   service.port,
                                                   service.protocol_port_str,
                                                   service.service_name_with_confidence,
                                                   service.service_confidence,
                                                   service.state_str,
                                                   service.nmap_service_state_reason,
                                                   service.nmap_product_version,
                                                   is_http,
                                                   url_str[0] if url_str else None,
                                                   tls_info.version_str,
                                                   tls_info.preference_str,
                                                   tls_info.heartbleed,
                                                   tls_info.compressors_str,
                                                   mapping.order,
                                                   cipher_suite.iana_name,
                                                   mapping.prefered,
                                                   kex_algorithm,
                                                   cipher_suite.security_str,
                                                   network_id,
                                                   host.id,
                                                   None,
                                                   service.id,
                                                   mapping.id,
                                                   network_sources,
                                                   host_sources,
                                                   None,
                                                   service.sources_str,
                                                   mapping.sources_str])
                for host_name in host_names:
                    environment = self._domain_config.get_environment(host_name)
                    host_name_sources = host_name.sources_str
                    network_str = host.ipv4_network.network if host.ipv4_network else None
                    for service in host_name.services:
                        if service.state in [ServiceState.Open, ServiceState.Closed] and \
                                descriptor.match_nmap_service_name(service):
                            url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"]
                            patch_missing_kex = {}
                            for tls_info in service.tls_info:
                                if self._filter(tls_info):
                                    for mapping in tls_info.cipher_suite_mappings:
                                        cipher_suite = mapping.cipher_suite
                                        key = "{}{}{}{}".format(host_name.full_name,
                                                                service.protocol_port_str,
                                                                tls_info.version_str,
                                                                cipher_suite.iana_name)
                                        kex_algorithm = mapping.kex_algorithm_details_str
                                        if kex_algorithm:
                                            patch_missing_kex[key] = kex_algorithm
                                        elif key in patch_missing_kex:
                                            kex_algorithm = patch_missing_kex[key]
                                        result.append([workspace.name,
                                                       "VHost",
                                                       network_str,
                                                       network_scope,
                                                       network_companies,
                                                       host.address,
                                                       host.version_str,
                                                       host_is_private,
                                                       host.in_scope,
                                                       host.os_family,
                                                       host.os_details,
                                                       host_name.domain_name.name,
                                                       host_name.domain_name.scope_str,
                                                       host_name.full_name,
                                                       host_name._in_scope,
                                                       host_name.name,
                                                       host_name.companies_str,
                                                       environment,
                                                       service.protocol_str,
                                                       service.port,
                                                       service.protocol_port_str,
                                                       service.service_name_with_confidence,
                                                       service.service_confidence,
                                                       service.state_str,
                                                       service.nmap_service_state_reason,
                                                       service.nmap_product_version,
                                                       True,
                                                       url_str[0] if url_str else None,
                                                       tls_info.version_str,
                                                       tls_info.preference_str,
                                                       tls_info.heartbleed,
                                                       tls_info.compressors_str,
                                                       mapping.order,
                                                       cipher_suite.iana_name,
                                                       mapping.prefered,
                                                       kex_algorithm,
                                                       cipher_suite.security_str,
                                                       network_id,
                                                       host.id,
                                                       host_name.id,
                                                       service.id,
                                                       mapping.id,
                                                       network_sources,
                                                       host_sources,
                                                       host_name_sources,
                                                       service.sources_str,
                                                       mapping.sources_str])
        return result

    def _final_report_tls_versions(self, workbook: Workbook):
        # Obtain results
        tls_results = {}
        versions = {}
        for workspace in self._workspaces:
            for host in workspace.hosts:
                if host.in_scope:
                    for service in host.services:
                        if service.tls_info:
                            summary = "{} {}".format(service.address, service.protocol_port_str)
                            if summary not in tls_results:
                                tls_results[summary] = {"service": service, "tls_versions": {}}
                            tls_entry = tls_results[summary]["tls_versions"]
                            for tls_info in service.tls_info:
                                tls_entry[tls_info.version_str] = True
                                versions[tls_info.version_str] = None
        if versions:
            # Convert results into two-dimensional array
            result = [["IP Address", "Service"]]
            if self._args.language == ReportLanguage.de:
                result = [["IP-Adresse", "Dienst"]]
            unique_tls_versions = list(versions.keys())
            unique_tls_versions.sort()
            result[0].extend(unique_tls_versions)
            for key, value in tls_results.items():
                service = value["service"]
                version = value["tls_versions"]
                row = [service.address, service.protocol_port_str]
                for item in unique_tls_versions:
                    if item in version:
                        row.append(self.TRUE)
                    else:
                        row.append(None)
                result.append(row)
            if len(result) > 1:
                self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                      csv_list=result,
                                      name="TLS - Versions",
                                      title="",
                                      description="")

    def _final_report_tls_ciphers(self, workbook: Workbook):
        result = [["IP Address", "Service", "Cipher Suite", "Security"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Cipher Suite", "Sicherheit"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                if host.in_scope:
                    for service in host.services:
                        for tls_info in service.tls_info:
                            for mapping in tls_info.cipher_suite_mappings:
                                if mapping.cipher_suite.security in [CipherSuiteSecurity.insecure,
                                                                     CipherSuiteSecurity.weak]:
                                    result.append([host.address,
                                                   service.protocol_port_str,
                                                   mapping.cipher_suite.iana_name,
                                                   mapping.cipher_suite.security_str])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="TLS - Weak Ciphers",
                                  title="",
                                  description="Classification of cipher suite security is coming from: "
                                              "https://ciphersuite.info")

    def final_report(self, workbook: Workbook):
        """
        This method creates all tables that are relevant to the final report.
        """
        self._final_report_tls_versions(workbook)
        self._final_report_tls_ciphers(workbook)


class _CertInfoReportGenerator(_BaseReportGenerator):
    """
    This class creates all reports for certificate information
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="cert info",
                         title="Overview Certificates",
                         description="The table provides an overview of all identified certificates.",
                         **kwargs)

    def _filter(self, cert_info: CertInfo) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return cert_info.is_processable(included_items=self._included_items,
                                        excluded_items=self._excluded_items,
                                        scope=self.scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        descriptor = HttpServiceDescriptor()
        result = [["Workspace",
                   "Type",
                   "Network (NW)",
                   "Scope (NW)",
                   "Company (NW)",
                   "IP Address (IP)",
                   "Version (IP)",
                   "Private IP",
                   "In Scope (IP)",
                   "OS Family",
                   "OS Details",
                   "Second-Level Domain (SLD)",
                   "Scope (SLD)",
                   "Host Name (HN)",
                   "In Scope (HN)",
                   "Name (HN)",
                   "Company (HN)",
                   "Environment",
                   "UDP/TCP",
                   "Port",
                   "Service (SRV)",
                   "Nmap Name (SRV)",
                   "Confidence (SRV)",
                   "State (SRV)",
                   "Reason State",
                   "Banner Information",
                   "Is HTTP",
                   "URL",
                   "HN Coverage",
                   "Common Name",
                   "Issuer",
                   "Invalid CA (Self-Signed)",
                   "Public Key Algorithm",
                   "Key Length",
                   "Public Key Summary",
                   "Proper Key Length",
                   "Hash Algorithm",
                   "Weak Signature",
                   "Cert. Type",
                   "Valid From",
                   "Valid Until",
                   "Valid Years",
                   "Is Valid",
                   "Within Recommended Validity",
                   "Subject Alternative Names",
                   "Key Usage",
                   "Critical Extensions",
                   "Serial Number",
                   "DB ID (NW)",
                   "DB ID (IP)",
                   "DB ID (HN)",
                   "DB ID (SRV)",
                   "DB ID (CERT)",
                   "Source (NW)",
                   "Source (IP)",
                   "Source (HN)",
                   "Source (SRV)",
                   "Source (CERT)"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                host_names = [mapping.host_name
                              for mapping in host.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                               DnsResourceRecordType.aaaa])]
                network_str = host.ipv4_network.network if host.ipv4_network else None
                network_id = host.ipv4_network.id if host.ipv4_network else None
                network_companies = host.ipv4_network.companies_str if host.ipv4_network else None
                network_sources = host.ipv4_network.sources_str if host.ipv4_network else None
                network_scope = host.ipv4_network.scope_str if host.ipv4_network else None
                host_is_private = host.ip_address.is_private
                host_sources = host.sources_str
                for service in host.services:
                    if service.state in [ServiceState.Open, ServiceState.Closed]:
                        is_http = descriptor.match_nmap_service_name(service)
                        url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"] \
                            if is_http else []
                        for cert_info in service.cert_info:
                            if self._filter(cert_info):
                                matching_host = "n/a"
                                result.append([workspace.name,
                                               "Host",
                                               network_str,
                                               network_scope,
                                               network_companies,
                                               host.address,
                                               host.version_str,
                                               host_is_private,
                                               host.in_scope,
                                               host.os_family,
                                               host.os_details,
                                               None,
                                               None,
                                               host.address,
                                               host.in_scope,
                                               None,
                                               None,
                                               None,
                                               service.protocol_str,
                                               service.port,
                                               service.protocol_port_str,
                                               service.service_name_with_confidence,
                                               service.service_confidence,
                                               service.state_str,
                                               service.nmap_service_state_reason,
                                               service.nmap_product_version,
                                               is_http,
                                               url_str[0] if url_str else None,
                                               matching_host,
                                               cert_info.common_name.lower(),
                                               cert_info.issuer_name.lower(),
                                               cert_info.is_self_signed(),
                                               cert_info.signature_asym_algorithm_str,
                                               cert_info.signature_bits,
                                               cert_info.signature_asym_algorithm_summary,
                                               None,
                                               cert_info.hash_algorithm_str,
                                               cert_info.has_weak_signature(),
                                               cert_info.cert_type_str,
                                               cert_info.valid_from_str,
                                               cert_info.valid_until_str,
                                               cert_info.validity_period_days / 365,
                                               cert_info.is_valid(),
                                               cert_info.has_recommended_duration(),
                                               cert_info.subject_alt_names_str.lower(),
                                               cert_info.key_usage_str,
                                               ", ".join(cert_info.critical_extension_names),
                                               cert_info.serial_number,
                                               network_id,
                                               host.id,
                                               None,
                                               service.id,
                                               cert_info.id,
                                               network_sources,
                                               host_sources,
                                               None,
                                               service.sources_str,
                                               cert_info.sources_str])
                for host_name in host_names:
                    environment = self._domain_config.get_environment(host_name)
                    host_name_sources = host_name.sources_str
                    network_str = host.ipv4_network.network if host.ipv4_network else None
                    for service in host_name.services:
                        if service.state in [ServiceState.Open, ServiceState.Closed] and \
                                descriptor.match_nmap_service_name(service):
                            is_http = descriptor.match_nmap_service_name(service)
                            url_str = [path.get_urlparse().geturl() for path in service.paths if path.name == "/"] \
                                if is_http else []
                            for cert_info in service.cert_info:
                                if self._filter(cert_info):
                                    matching_host = cert_info.matches_host_name(host_name) \
                                        if cert_info.cert_type == CertType.identity else None
                                    result.append([workspace.name,
                                                   "VHost",
                                                   network_str,
                                                   network_scope,
                                                   network_companies,
                                                   host.address,
                                                   host.version_str,
                                                   host_is_private,
                                                   host.in_scope,
                                                   host.os_family,
                                                   host.os_details,
                                                   host_name.domain_name.name,
                                                   host_name.domain_name.scope_str,
                                                   host_name.full_name,
                                                   host_name._in_scope,
                                                   host_name.name,
                                                   host_name.companies_str,
                                                   environment,
                                                   service.protocol_str,
                                                   service.port,
                                                   service.protocol_port_str,
                                                   service.service_name_with_confidence,
                                                   service.service_confidence,
                                                   service.state_str,
                                                   service.nmap_service_state_reason,
                                                   service.nmap_product_version,
                                                   True,
                                                   url_str[0] if url_str else None,
                                                   matching_host,
                                                   cert_info.common_name.lower(),
                                                   cert_info.issuer_name.lower(),
                                                   cert_info.is_self_signed(),
                                                   cert_info.signature_asym_algorithm_str,
                                                   cert_info.signature_bits,
                                                   cert_info.signature_asym_algorithm_summary,
                                                   None,
                                                   cert_info.hash_algorithm_str,
                                                   cert_info.has_weak_signature(),
                                                   cert_info.cert_type_str,
                                                   cert_info.valid_from_str,
                                                   cert_info.valid_until_str,
                                                   cert_info.validity_period_days / 365,
                                                   cert_info.is_valid(),
                                                   cert_info.has_recommended_duration(),
                                                   cert_info.subject_alt_names_str.lower(),
                                                   cert_info.key_usage_str,
                                                   ", ".join(cert_info.critical_extension_names),
                                                   cert_info.serial_number,
                                                   network_id,
                                                   host.id,
                                                   host_name.id,
                                                   service.id,
                                                   cert_info.id,
                                                   network_sources,
                                                   host_sources,
                                                   host_name_sources,
                                                   service.sources_str,
                                                   cert_info.sources_str])
        return result

    def _final_report_host_name_coverage(self, workbook: Workbook):
        result = [["IP Address", "Service", "Host Names", "Common Name", "Subject Alternative\nNames", "Full\nCoverage"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Hostnamen", "Common Name", "Subject Alternative\nNames", "Volle\nAbdeckung"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                host_names = [item.host_name.full_name
                              for item in host.get_host_host_name_mappings([DnsResourceRecordType.a,
                                                                            DnsResourceRecordType.aaaa])]
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            matching_host = cert_info.matches_host_names(host_names)
                            result.append([host.address,
                                           service.protocol_port_str,
                                           ", ".join(host_names),
                                           cert_info.common_name,
                                           cert_info.subject_alt_names_str,
                                           self.TRUE if matching_host else None])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Name Coverage",
                                  title="",
                                  description="")

    def _final_report_valid_ca(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Common Name", "Issuer", "Self-\nSigned"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Common Name", "Issuer", "Selbst\nSigniert"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           cert_info.common_name,
                                           cert_info.issuer_name,
                                           self.TRUE if cert_info.is_self_signed() else None])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Valid CAs",
                                  title="",
                                  description="")

    def _final_report_signature_algorithm(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Public Key\nAlgorithm", "Hash Algorithm"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Public-Key-\nAlgorithmus", "Hashalgorithmus"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           cert_info.signature_asym_algorithm_summary,
                                           cert_info.hash_algorithm_str])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Signing Algorithms",
                                  title="",
                                  description="")

    def _final_report_durations(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Valid\nFrom", "Valid\nUntil", "Valid", "Years", "Valid\nDuration"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Gültig\nvon", "Gültig\nbis", "Gültig", "Jahre", "Gültige\nDauer"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name,
                                           cert_info.valid_from_str,
                                           cert_info.valid_until_str,
                                           self.TRUE if cert_info.is_valid() else None,
                                           "{:.2f}".format(cert_info.validity_period_days / 365),
                                           self.TRUE if cert_info.has_recommended_duration() else None])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Durations",
                                  title="",
                                  description="")

    def _final_report_key_usages(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Key Usage"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Verwendungszweck"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           cert_info.key_usage_str])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Key Usage",
                                  title="",
                                  description="")

    def _final_report_critical_extensions(self, workbook: Workbook):
        result = [["IP Address", "Service", "Name", "Critical Extensions"]]
        if self._args.language == ReportLanguage.de:
            result = [["IP-Adresse", "Dienst", "Name", "Kritische Erweiterungen"]]
        for workspace in self._workspaces:
            for host in workspace.hosts:
                for service in host.services:
                    for cert_info in service.cert_info:
                        if cert_info.cert_type == CertType.identity:
                            result.append([host.address,
                                           service.protocol_port_str,
                                           service.service_name_with_confidence,
                                           ", ".join(cert_info.critical_extension_names)])
        if len(result) > 1:
            self.fill_excel_sheet(worksheet=workbook.create_sheet(),
                                  csv_list=result,
                                  name="Cert - Crit. Extensions",
                                  title="",
                                  description="")

    def final_report(self, workbook: Workbook):
        """
        This method creates all tables that are relevant to the final report.
        """
        self._final_report_host_name_coverage(workbook)
        self._final_report_valid_ca(workbook)
        self._final_report_signature_algorithm(workbook)
        self._final_report_durations(workbook)
        self._final_report_key_usages(workbook)
        self._final_report_critical_extensions(workbook)


class _CredentialReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for credentials
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="credential info",
                         title="Overview Identified Credentials",
                         description="The table provides an overview of all identified credentials.",
                         **kwargs)

    def _filter(self, credentials: Credentials) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return credentials.is_processable(included_items=self._included_items,
                                          excluded_items=self._excluded_items,
                                          scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID",
                 "Workspace",
                 "Type",
                 "Network",
                 "In Scope",
                 "Address",
                 "Address Summary",
                 "Host Names",
                 "Service Summary",
                 "TCP/UDP",
                 "Port",
                 "Service",
                 "Type",
                 "Complete",
                 "User name",
                 "Password",
                 "Source"]]
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    for service in host_name.services:
                        for credential in service.credentials:
                            if self._filter(credential):
                                rows.append([credential.id,
                                             workspace.name,
                                             "vhost",
                                             None,
                                             host_name.in_scope(CollectorType.vhost_service),
                                             service.address,
                                             service.address_summary,
                                             None,
                                             service.summary,
                                             service.protocol_str,
                                             service.port,
                                             service.protocol_port_str,
                                             credential.type_str,
                                             credential.complete,
                                             credential.username,
                                             credential.password,
                                             credential.sources_str])
            for host in workspace.hosts:
                host_names = host.get_host_host_name_mappings_str([DnsResourceRecordType.a,
                                                                   DnsResourceRecordType.aaaa,
                                                                   DnsResourceRecordType.ptr])
                if host.ipv4_network:
                    ipv4_network = host.ipv4_network.network
                else:
                    ipv4_network = None
                for service in host.services:
                    for credential in service.credentials:
                        if self._filter(credential):
                            rows.append([credential.id,
                                         workspace.name,
                                         "host",
                                         ipv4_network,
                                         host.in_scope,
                                         service.address,
                                         service.address_summary,
                                         host_names,
                                         service.summary,
                                         service.protocol_str,
                                         service.port,
                                         service.protocol_port_str,
                                         credential.type_str,
                                         credential.complete,
                                         credential.username,
                                         credential.password,
                                         credential.sources_str])
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    for email in host_name.emails:
                        for credential in email.credentials:
                            if self._filter(credential):
                                rows.append([credential.id,
                                             workspace.name,
                                             "email",
                                             None,
                                             domain.in_scope,
                                             None,
                                             None,
                                             None,
                                             None,
                                             None,
                                             None,
                                             None,
                                             credential.type_str,
                                             credential.complete,
                                             email.email_address,
                                             credential.password,
                                             credential.sources_str])
        return rows


class _AdditionalInfoReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for additional info
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="additional info",
                         title="Overview Identified Additional Information",
                         description="The table provides an overview of all identified additional information like "
                                     "version information. Columns 'Name' and 'Value' are key value pairs and you "
                                     "can sort column 'Name' to determine which values exist for certain keys.",
                         **kwargs)

    def _filter(self, additional_info: AdditionalInfo) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return additional_info.is_processable(included_items=self._included_items,
                                              excluded_items=self._excluded_items,
                                              scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID",
                 "Workspace",
                 "Type",
                 "Network",
                 "In Scope",
                 "Address",
                 "Address Summary",
                 "Resolves to",
                 "Service Summary",
                 "TCP/UDP",
                 "Port",
                 "Service",
                 "Name",
                 "Value",
                 "Sources"]]
        for workspace in self._workspaces:
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    ipv4_addresses = host_name.get_host_host_name_mappings_str(types=[DnsResourceRecordType.a,
                                                                                      DnsResourceRecordType.aaaa,
                                                                                      DnsResourceRecordType.ptr])
                    for service in host_name.services:
                        for additional_info in service.additional_info:
                            if additional_info.name not in ["CVEs",
                                                            HaveIBeenPwnedBreachedAcccount.NAME,
                                                            HaveIBeenPwnedPasteAcccount.NAME] and \
                                    self._filter(additional_info):
                                for value in additional_info.values:
                                    rows.append([additional_info.id,
                                                 workspace.name,
                                                 "vhost",
                                                 None,
                                                 host_name.in_scope(CollectorType.vhost_service),
                                                 service.address,
                                                 service.address_summary,
                                                 ipv4_addresses,
                                                 service.summary,
                                                 service.protocol_str,
                                                 service.port,
                                                 service.protocol_port_str,
                                                 additional_info.name,
                                                 value,
                                                 additional_info.sources_str])
            for host in workspace.hosts:
                host_names = host.get_host_host_name_mappings_str(types=[DnsResourceRecordType.a,
                                                                         DnsResourceRecordType.aaaa,
                                                                         DnsResourceRecordType.ptr])
                if host.ipv4_network:
                    ipv4_network = host.ipv4_network.network
                else:
                    ipv4_network = None
                for service in host.services:
                    for additional_info in service.additional_info:
                        if additional_info.name not in ["CVEs",
                                                        HaveIBeenPwnedBreachedAcccount.NAME,
                                                        HaveIBeenPwnedPasteAcccount.NAME] and \
                                self._filter(additional_info):
                            sources = additional_info.sources_str
                            for value in additional_info.values:
                                rows.append([additional_info.id,
                                             workspace.name,
                                             "host",
                                             ipv4_network,
                                             host.in_scope,
                                             service.address,
                                             service.address_summary,
                                             host_names,
                                             service.summary,
                                             service.protocol_str,
                                             service.port,
                                             service.protocol_port_str,
                                             additional_info.name,
                                             value,
                                             sources])
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    for additional_info in host_name.additional_info:
                        if additional_info.name not in ["CVEs",
                                                        HaveIBeenPwnedBreachedAcccount.NAME,
                                                        HaveIBeenPwnedPasteAcccount.NAME] \
                                and self._filter(additional_info):
                            sources = additional_info.sources_str
                            for value in additional_info.values:
                                rows.append([additional_info.id,
                                             workspace.name,
                                             "domain",
                                             None,
                                             host_name.in_scope(CollectorType.vhost_service),
                                             host_name.full_name,
                                             None,
                                             None,
                                             None,
                                             None,
                                             None,
                                             None,
                                             additional_info.name,
                                             value,
                                             sources])
            for domain in workspace.domain_names:
                for host_name in domain.host_names:
                    for email in host_name.emails:
                        for additional_info in email.additional_info:
                            if additional_info.name not in ["CVEs",
                                                            HaveIBeenPwnedBreachedAcccount.NAME,
                                                            HaveIBeenPwnedPasteAcccount.NAME] \
                                    and self._filter(additional_info):
                                sources = additional_info.sources_str
                                for value in additional_info.values:
                                    rows.append([additional_info.id,
                                                 workspace.name,
                                                 "email",
                                                 None,
                                                 domain.in_scope,
                                                 email.email_address,
                                                 None,
                                                 None,
                                                 None,
                                                 None,
                                                 None,
                                                 None,
                                                 additional_info.name,
                                                 value,
                                                 sources])
            for ipv4_network in workspace.ipv4_networks:
                for additional_info in ipv4_network.additional_info:
                    if additional_info.name not in ["CVEs",
                                                    HaveIBeenPwnedBreachedAcccount.NAME,
                                                    HaveIBeenPwnedPasteAcccount.NAME] and self._filter(additional_info):
                        sources = additional_info.sources_str
                        for value in additional_info.values:
                            rows.append([additional_info.id,
                                         workspace.name,
                                         "network",
                                         ipv4_network.network,
                                         ipv4_network.in_scope,
                                         ipv4_network.network,
                                         None,
                                         None,
                                         None,
                                         None,
                                         None,
                                         None,
                                         additional_info.name,
                                         value,
                                         sources])
            for company in workspace.companies:
                for additional_info in company.additional_info:
                    sources = additional_info.sources_str
                    for value in additional_info.values:
                        rows.append([additional_info.id,
                                     workspace.name,
                                     "company",
                                     None,
                                     company.in_scope,
                                     company.name,
                                     None,
                                     None,
                                     None,
                                     None,
                                     None,
                                     None,
                                     additional_info.name,
                                     value,
                                     sources])
        return rows


class _NetworkReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for networks
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="network info",
                         title="Overview Identified Networks",
                         description="The table provides an overview of all identified networks. Note that the column "
                                     "'Number IPs' contains the number of hosts that were identified and are "
                                     "associated with this network. In other words, this column provides an indicator "
                                     "how extensive the identified network ranges are used.",
                         **kwargs)

    def _filter(self, network: Network) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return network.is_processable(included_items=self._included_items,
                                      excluded_items=self._excluded_items,
                                      scope=self._scope)

    def _egrep_text(self, ipv4_network: Network) -> List[str]:
        """
        This method returns all lines matching the given list of regular expressions
        :param ipv4_network: The network whose text output shall be parsed
        :return:
        """
        result = self._egrep(ipv4_network.get_text(ident=0,
                                                   exclude_collectors=self._excluded_collectors,
                                                   include_collectors=self._included_collectors,
                                                   scope=self._scope,
                                                   show_metadata=False))
        return result

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID", "Workspace", "Network", "Version", "Companies", "Scope", "Sources", "Number IPs"]]
        for workspace in self._workspaces:
            for network in workspace.ipv4_networks:
                if self._filter(network):
                    rows.append([network.id,
                                 workspace.name,
                                 network.network,
                                 network.version_str,
                                 network.companies_str,
                                 network.scope_str,
                                 network.sources_str,
                                 len(network.hosts)])
        return rows

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        for workspace in self._workspaces:
            for ipv4_network in workspace.ipv4_networks:
                if self._filter(ipv4_network):
                    rvalue.extend(ipv4_network.get_text(ident=0,
                                                        scope=self.scope,
                                                        exclude_collectors=self._excluded_collectors,
                                                        include_collectors=self._included_collectors,
                                                        report_visibility=self._visibility,
                                                        color=self._color))
        return rvalue

    def grep_text(self) -> List[List[str]]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rows = [["DB ID", "Workspace", "Network", "Companies", "In Scope", "Result"]]
        for workspace in self._workspaces:
            for ipv4_network in workspace.ipv4_networks:
                if self._filter(ipv4_network):
                    results = self._egrep_text(ipv4_network)
                    if self._not_grep and not results:
                        rows.append([ipv4_network.id,
                                     workspace.name,
                                     ipv4_network.network,
                                     ipv4_network.companies_str,
                                     ipv4_network.in_scope,
                                     None])
                    elif not self._not_grep:
                        for row in results:
                            rows.append([ipv4_network.id,
                                         workspace.name,
                                         ipv4_network.network,
                                         ipv4_network.companies_str,
                                         ipv4_network.in_scope,
                                         row])
        return rows


class _EmailReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for emails
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="email info",
                         title="Overview Identified Email Addresses",
                         description="The table provides an overview of all identified email addresses.",
                         **kwargs)

    def _filter(self, email: Email) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return email.is_processable(included_items=self._included_items,
                                    excluded_items=self._excluded_items,
                                    scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID", "Workspace", "Email", "Domain", "In Scope", "Sources"]]
        for workspace in self._workspaces:
            email_addresses = self._session.query(Email)\
                .join(HostName)\
                .join(DomainName)\
                .join(Workspace)\
                .filter(Workspace.id == workspace.id).all()
            for email_address in email_addresses:
                if self._filter(email_address):
                    row = [email_address.id,
                           workspace.name,
                           email_address.email_address,
                           email_address.host_name.full_name,
                           email_address.host_name.domain_name.in_scope,
                           email_address.sources_str]
                    rows.append(row)
        return rows

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        for workspace in self._workspaces:
            email_addresses = self._session.query(Email) \
                .join(HostName) \
                .join(DomainName) \
                .join(Workspace) \
                .filter(Workspace.id == workspace.id).all()
            for email in email_addresses:
                if self._filter(email):
                    rvalue.extend(email.get_text(ident=0,
                                                 scope=self.scope,
                                                 report_visibility=self._visibility,
                                                 exclude_collectors=self._excluded_collectors,
                                                 include_collectors=self._included_collectors,
                                                 color=self._color))
        return rvalue


class _CompanyReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for company
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="company info",
                         title="Overview Identified Company Names",
                         description="The table provides an overview of all identified companies.",
                         **kwargs)

    def _filter(self, company: Company) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return company.is_processable(included_items=self._included_items,
                                      excluded_items=self._excluded_items,
                                      scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID", "Workspace", "Company", "Owns", "Owns Type", "In Scope", "Owns Scope", "Sources"]]
        for workspace in self._workspaces:
            results = self._session.query(Company)\
                .join(Workspace)\
                .filter(Workspace.id == workspace.id).all()
            for company in results:
                if self._filter(company):
                    has_results = False
                    in_scope = company.in_scope
                    sources = company.sources_str
                    for network in company.networks:
                        has_results = True
                        row = [company.id,
                               company.workspace.name,
                               company.name,
                               network.network,
                               "network",
                               in_scope,
                               network.scope_str,
                               sources]
                        rows.append(row)
                    for domain in company.domain_names:
                        has_results = True
                        row = [company.id,
                               company.workspace.name,
                               company.name,
                               domain.name,
                               "domain",
                               in_scope,
                               domain.scope_str,
                               sources]
                        rows.append(row)
                    if not has_results:
                        row = [company.id,
                               company.workspace.name,
                               company.name,
                               None,
                               None,
                               in_scope,
                               None,
                               sources]
                        rows.append(row)
        return rows

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        for workspace in self._workspaces:
            results = self._session.query(Company) \
                .join(Workspace) \
                .filter(Workspace.id == workspace.id).all()
            for company in results:
                if self._filter(company):
                    rvalue.extend(company.get_text(ident=0,
                                                   scope=self.scope,
                                                   report_visibility=self._visibility,
                                                   exclude_collectors=self._excluded_collectors,
                                                   include_collectors=self._included_collectors,
                                                   color=self._color))
        return rvalue


class _BreachReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for emails
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="breach info",
                         title="Overview Identified Breaches",
                         description="The table provides an overview of all breaches collected from "
                                     "haveibeenpwned.com.",
                         **kwargs)

    def _filter(self, email: Email) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return email.is_processable(included_items=self._included_items,
                                    excluded_items=self._excluded_items,
                                    scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID", "Workspace", "Email", "Host Name", "In Scope", "Breached", "Type", "Value", "Sources"]]
        for workspace in self._workspaces:
            email_addresses = self._session.query(Email)\
                .join(HostName)\
                .join(DomainName)\
                .join(Workspace)\
                .filter(Workspace.id == workspace.id).all()
            for email_address in email_addresses:
                sources = email_address.sources_str
                if self._filter(email_address):
                    additional_info = self._session.query(AdditionalInfo) \
                        .join(Email) \
                        .join(HostName) \
                        .join(DomainName) \
                        .join(Workspace) \
                        .filter(Workspace.id == workspace.id,
                                Email.id == email_address.id).all()
                    for info in additional_info:
                        if info.name in [HaveIBeenPwnedBreachedAcccount.NAME, HaveIBeenPwnedPasteAcccount.NAME]:
                            for item in info.values:
                                row = [email_address.id,
                                       workspace.name,
                                       email_address.email_address,
                                       email_address.host_name.full_name,
                                       email_address.host_name.domain_name.in_scope,
                                       True,
                                       info.name,
                                       item,
                                       sources]
                                rows.append(row)
                    if len(additional_info) == 0:
                        row = [email_address.id,
                               workspace.name,
                               email_address.email_address,
                               email_address.host_name.full_name,
                               email_address.host_name.domain_name.in_scope,
                               False,
                               None,
                               None,
                               sources]
                        rows.append(row)
        return rows

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        for workspace in self._workspaces:
            email_addresses = self._session.query(Email) \
                .join(HostName) \
                .join(DomainName) \
                .join(Workspace) \
                .filter(Workspace.id == workspace.id).all()
            for email in email_addresses:
                if self._filter(email):
                    rvalue.extend(email.get_text(ident=0,
                                                 scope=self.scope,
                                                 report_visibility=self._visibility,
                                                 exclude_collectors=self._excluded_collectors,
                                                 include_collectors=self._included_collectors,
                                                 color=self._color))
        return rvalue


class _FileReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for emails
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="file info",
                         title="Overview of Collected Files",
                         description="The table provides an overview of all collected files (e.g., certificates or raw "
                                     "outputs of certain collectors). Note that file types (see column 'File Type') "
                                     "text, json, and xml are usually the raw outputs of the executed operating "
                                     "systems commands. File type screenshot represents screenshots of web application "
                                     "that were created by the tool eyewitness. Finally, file type certificate are "
                                     "the raw certificates files (e.g., PEM format) that were collected by tools like"
                                     "Nmap, Sslyze, OpenSSL, or Sslscan. The command kisreport allows exporting these "
                                     "files via the positional argument file.",
                         **kwargs)
        if "type" in args:
            if "all" in args.type:
                self._file_types = [FileType[item.name] for item in FileType]
            else:
                self._file_types = [FileType[item] for item in args.type]
        else:
            self._file_types = []

    def _filter(self, command: Command) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return command.is_processable(included_items=self._included_items,
                                      excluded_items=self._excluded_items,
                                      exclude_collectors=self._excluded_collectors,
                                      include_collectors=self._included_collectors,
                                      scope=self._scope)

    def _export_file(self, file_info: CommandFileMapping, deduplicated: Dict[str, str] = {}):
        """
        This method writes the given file_info to the filesystem
        :param file_info:
        :param deduplicated:
        :return:
        """
        deduplicated[file_info.file.sha256_value] = True
        file_name = self._get_unique_file_name(self._args.export_path, file_info.file_name)
        if os.path.isfile(file_name):
            print("file '{}' exists already but will be overwritten.".format(file_name),
                  file=sys.stderr)
        with open(file_name, "wb") as file:
            file.write(file_info.file.content)

    def _append_csv_row(self, csv_rows: List[List[str]], command: Command, file_type: FileType) -> None:
        service = command.service
        if not self._file_types or file_type in self._file_types:
            stdout = command.stdout
            content_length = None
            if file_type == FileType.xml and command.xml_output:
                content_length = len(str(command.xml_output))
            elif file_type == FileType.json and command.json_output:
                content_length = len(str(command.json_output))
            elif file_type == FileType.binary and command.binary_output:
                content_length = len(command.binary_output)
            elif file_type == FileType.text and stdout:
                content_length = len(os.linesep.join(stdout))
            if content_length:
                csv_rows.append([command.id,
                                 command.workspace.name,
                                 command.file_name,
                                 file_type.name.lower(),
                                 content_length,
                                 command.collector_name.name,
                                 command.collector_name.type_str,
                                 command.target_name,
                                 service.port if service else None,
                                 service.protocol_str if service else None,
                                 service.service_name if service else None,
                                 command.status_str])

    def _export_raw_scan_result(self, command: Command, file_type: FileType) -> None:
        """
        This method writes the raw scan results to the filesystem
        :param command:
        :param file_type:
        :return:
        """
        if not self._file_types or file_type in self._file_types:
            contents = []
            file_name = None
            stdout = command.stdout
            if file_type == FileType.xml and command.xml_output:
                file_name = "{}.xml".format(command.file_name)
                contents = [command.xml_output]
            elif file_type == FileType.json and command.json_output:
                file_name = "{}.json".format(command.file_name)
                contents = [json.dumps(item, indent=4) for item in command.json_output if item]
            elif file_type == FileType.binary and command.binary_output:
                file_name = "{}.bin".format(command.file_name)
                contents = [command.binary_output]
            elif file_type == FileType.text and stdout:
                file_name = "{}.txt".format(command.file_name)
                contents = [os.linesep.join(stdout)]
            for content in contents:
                file_path = self._get_unique_file_name(self._args.export_path, file_name)
                mode = "w" if isinstance(content, str) else "wb"
                with open(file_path, mode) as file:
                    file.write(content)

    def export_files(self) -> None:
        """
        Exports all files from the database.
        :return:
        """
        commands = self._session.query(Command)
        deduplicated = {}
        for command in commands.all():
            if command.workspace in self._workspaces:
                if self._filter(command):
                    self._export_raw_scan_result(command, FileType.text)
                    self._export_raw_scan_result(command, FileType.xml)
                    self._export_raw_scan_result(command, FileType.json)
                    self._export_raw_scan_result(command, FileType.binary)
                    for mapping in command.file_mappings:
                        if not self._file_types or mapping.file.type in self._file_types:
                            self._export_file(mapping, deduplicated)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID",
                 "Workspace",
                 "File Name",
                 "File Type",
                 "File Size (bytes)",
                 "Collector",
                 "Collector Type",
                 "Address",
                 "Protocol",
                 "Port",
                 "Service Name",
                 "Status"]]
        commands = self._session.query(Command)
        for command in commands:
            if command.workspace in self._workspaces:
                if self._filter(command):
                    self._append_csv_row(rows, command, FileType.text)
                    self._append_csv_row(rows, command, FileType.xml)
                    self._append_csv_row(rows, command, FileType.json)
                    self._append_csv_row(rows, command, FileType.binary)
                    for mapping in command.file_mappings:
                        if not self._file_types or mapping.file.type in self._file_types:
                            service = command.service
                            rows.append([command.id,
                                         command.workspace.name,
                                         mapping.file_name,
                                         mapping.file.type_str,
                                         len(mapping.file.content),
                                         command.collector_name.name,
                                         command.collector_name.type_str,
                                         command.target_name,
                                         service.port if service else None,
                                         service.protocol_str if service else None,
                                         service.service_name if service else None,
                                         command.status_str])
        return rows


class _CollectorReportGenerator(_BaseReportGenerator):
    """
    This method creates all reports for collectors
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="command info",
                         title="Overview of Executed Operating System Commands",
                         description="The table provides an overview of all executed commands. You can use columns "
                                     "'Start Time [UTC]' and 'End Time [UTC]' to determine when a certain command was "
                                     "executed. You can use 'Stdout Size' or 'Stderr Size' to determine deviations "
                                     "from the average command output of a specific collector.",
                         **kwargs)

    def _filter(self, command: Command) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return command.is_processable(included_items=self._included_items,
                                      excluded_items=self._excluded_items,
                                      exclude_collectors=self._excluded_collectors,
                                      include_collectors=self._included_collectors,
                                      scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["DB ID",
                 "Workspace",
                 "Collector",
                 "Type",
                 "Address",
                 "Protocol",
                 "Port",
                 "Service",
                 "Nmap Service Name",
                 "Nmap Service Name Original",
                 "Status",
                 "Start Time [UTC]",
                 "End Time [UTC]",
                 "Duration [s]",
                 "Return Code",
                 "Stdout Size",
                 "Stderr Size",
                 "OS Command"]]
        commands = self._session.query(Command)
        for command in commands:
            service = command.service
            if command.workspace in self._workspaces:
                if self._filter(command):
                    execution_time = (command.stop_time - command.start_time).seconds \
                        if command.stop_time and command.start_time else None
                    start_time = command.start_time_str
                    stop_time = command.stop_time_str
                    stdout_count = len(os.linesep.join(command.stdout_output))
                    stderr_count = len(os.linesep.join(command.stderr_output))
                    rows.append([command.id,
                                 command.workspace.name,
                                 command.collector_name.name,
                                 command.collector_name.type_str,
                                 command.target_name,
                                 service.protocol_str if service else None,
                                 service.port if service else None,
                                 service.protocol_port_str if service else None,
                                 service.service_name_with_confidence if service else None,
                                 service.nmap_service_name_original_with_confidence if service else None,
                                 command.status_str,
                                 start_time,
                                 stop_time,
                                 execution_time,
                                 command.return_code,
                                 stdout_count,
                                 stderr_count,
                                 command.os_command_string])
        return rows

    def get_text(self) -> List[str]:
        """
        This method returns all information as a list of text.
        :return:
        """
        rvalue = []
        commands = self._session.query(Command)
        for command in commands:
            if command.workspace in self._workspaces:
                if self._filter(command):
                    rvalue += command.get_text(ident=0,
                                               report_visibility=self._visibility,
                                               color=self._color)
        return rvalue


class _VulnerabilityReportGenerator(_BaseReportGenerator):
    """
    This method creates report for all vulnerabilities
    """

    def __init__(self, args, session: Session, workspaces: List[Workspace], **kwargs) -> None:
        super().__init__(args,
                         session,
                         workspaces,
                         name="vulnerabilities",
                         title="Overview of Identified Vulnerabilities",
                         description="The table provides an overview of all vulnerabilities, which were identified by "
                                     "Nessus or Shodan.io.",
                         **kwargs)

    def _filter(self, additional_info: AdditionalInfo) -> bool:
        """
        Method determines whether the given item shall be included into the report
        """
        return additional_info.is_processable(included_items=self._included_items,
                                              excluded_items=self._excluded_items,
                                              scope=self._scope)

    def get_csv(self) -> List[List[str]]:
        """
        This method returns all information as CSV.
        :return:
        """
        rows = [["Workspace",
                 "Network (NW)",
                 "Scope (NW)",
                 "Company (NW)",
                 "IP Address (IP)",
                 "IP Summary",
                 "In Scope (IP)",
                 "Host Names (HN)",
                 "Protocol",
                 "Port",
                 "Service",
                 "Nmap Service Name",
                 "Nessus Service Name",
                 "Sources",
                 "CVE",
                 "CVSSv3",
                 "CVSSv2",
                 "Plugin ID",
                 "Description"]]
        additional_info = self._session.query(AdditionalInfo)\
            .join(Service)\
            .join(Host)\
            .join(Workspace)\
            .filter(Workspace.name.in_([item.name for item in self._workspaces]),
                    AdditionalInfo.name == "CVEs").all()
        for item in additional_info:
            if self._filter(item):
                network = None
                companies = None
                network_scope = None
                host = item.service.host
                host_names = host.get_host_host_name_mappings_str([DnsResourceRecordType.a,
                                                                   DnsResourceRecordType.aaaa,
                                                                   DnsResourceRecordType.ptr])
                if host.ipv4_network:
                    network = host.ipv4_network.network
                    companies = host.ipv4_network.companies_str
                    network_scope = host.ipv4_network.scope_str
                for entry in BaseUtils.get_csv_as_list(item.values):
                    tmp = [host.workspace.name,
                           network,
                           network_scope,
                           companies,
                           host.address,
                           host.summary,
                           host.in_scope,
                           host_names,
                           item.service.protocol_str,
                           item.service.port,
                           item.service.protocol_port_str,
                           item.service.nmap_service_name,
                           item.service.nessus_service_name,
                           item.sources_str]
                    tmp.extend(entry)
                    rows.append(tmp)
        return rows


class ReportGenerator:
    """This class creates all reports"""

    def __init__(self,
                 args,
                 session: Session,
                 workspaces: List[Workspace], **kwargs):
        self._generators = {ExcelReport.host.name: _HostReportGenerator,
                            ExcelReport.vhost.name: _HostNameReportGenerator,
                            ExcelReport.domain.name: _DomainNameReportGenerator,
                            ExcelReport.cname.name: _CanonicalNameReportGenerator,
                            ExcelReport.network.name: _NetworkReportGenerator,
                            ExcelReport.path.name: _PathReportGenerator,
                            ExcelReport.credential.name: _CredentialReportGenerator,
                            ExcelReport.email.name: _EmailReportGenerator,
                            ExcelReport.company.name: _CompanyReportGenerator,
                            ExcelReport.additionalinfo.name: _AdditionalInfoReportGenerator,
                            ExcelReport.file.name: _FileReportGenerator,
                            ExcelReport.breach.name: _BreachReportGenerator,
                            ExcelReport.vulnerability.name: _VulnerabilityReportGenerator,
                            ExcelReport.command.name: _CollectorReportGenerator,
                            ExcelReport.tls.name: _TlsInfoReportGenerator,
                            ExcelReport.cert.name: _CertInfoReportGenerator}
        self._args = args
        self._workspaces = workspaces
        self._session = session

    def run(self) -> None:
        """
        This method runs the desired report
        :return:
        """
        module = self._args.module.replace("-", "")
        if module in self._generators:
            self._generators[module](self._args, self._session, self._workspaces).export()
        elif self._args.module == "excel":
            if os.path.exists(self._args.FILE):
                os.unlink(self._args.FILE)
            workbook = Workbook()
            first = True
            for report_str in self._args.reports:
                print("* creating report for: {}".format(report_str))
                generator = self._generators[report_str]
                instance = generator(self._args, self._session, self._workspaces)
                csv_list = instance.get_csv()
                if len(csv_list) > 1:
                    if first:
                        instance.fill_excel_sheet(workbook.active, csv_list=csv_list)
                        first = False
                    else:
                        instance.fill_excel_sheet(workbook.create_sheet(), csv_list=csv_list)
            workbook.save(self._args.FILE)
        elif self._args.module == "final":
            if os.path.exists(self._args.FILE):
                os.unlink(self._args.FILE)
            workbook = Workbook()
            workbook.remove(workbook.active)
            for generator in self._generators.values():
                instance = generator(self._args, self._session, self._workspaces)
                instance.final_report(workbook=workbook)
            workbook.save(self._args.FILE)
